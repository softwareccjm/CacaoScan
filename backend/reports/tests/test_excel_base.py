"""
Tests for ExcelBaseGenerator.
"""
import pytest
from unittest.mock import Mock, patch, MagicMock
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.utils import timezone
from django.contrib.auth.models import User

from reports.services.report.excel.excel_base import ExcelBaseGenerator


@pytest.fixture
def excel_generator():
    """Create ExcelBaseGenerator instance."""
    return ExcelBaseGenerator()


@pytest.fixture
def user(db):
    """Create test user with unique username and email."""
    import uuid
    unique_id = str(uuid.uuid4())[:8]
    return User.objects.create_user(
        username=f'testuser_{unique_id}',
        email=f'test_{unique_id}@example.com',
        password='testpass123',
        first_name='Test',
        last_name='User'
    )


class TestExcelBaseGenerator:
    """Tests for ExcelBaseGenerator class."""
    
    def test_init(self, excel_generator):
        """Test generator initialization."""
        assert excel_generator.workbook is None
        assert excel_generator.ws is None
    
    def test_get_thin_border(self, excel_generator):
        """Test _get_thin_border method."""
        border = excel_generator._get_thin_border()
        assert isinstance(border, Border)
        assert border.left.style == 'thin'
        assert border.right.style == 'thin'
        assert border.top.style == 'thin'
        assert border.bottom.style == 'thin'
    
    def test_get_header_fill(self, excel_generator):
        """Test _get_header_fill method."""
        fill = excel_generator._get_header_fill()
        assert isinstance(fill, PatternFill)
        assert fill.start_color.rgb[2:] == ExcelBaseGenerator.HEADER_COLOR
        assert fill.end_color.rgb[2:] == ExcelBaseGenerator.HEADER_COLOR
        assert fill.fill_type == "solid"
    
    def test_get_header_font(self, excel_generator):
        """Test _get_header_font method."""
        font = excel_generator._get_header_font()
        assert isinstance(font, Font)
        assert font.bold is True
        assert font.color.rgb[2:] == ExcelBaseGenerator.HEADER_TEXT_COLOR
    
    def test_get_alignment_map(self, excel_generator):
        """Test _get_alignment_map method."""
        alignment_map = excel_generator._get_alignment_map()
        assert 'center' in alignment_map
        assert 'left' in alignment_map
        assert 'right' in alignment_map
        assert isinstance(alignment_map['center'], Alignment)
        assert isinstance(alignment_map['left'], Alignment)
        assert isinstance(alignment_map['right'], Alignment)
    
    def test_create_workbook(self, excel_generator):
        """Test _create_workbook method."""
        excel_generator._create_workbook("Test Sheet")
        assert excel_generator.workbook is not None
        assert excel_generator.ws is not None
        assert excel_generator.ws.title == "Test Sheet"
    
    def test_create_workbook_default_title(self, excel_generator):
        """Test _create_workbook with default title."""
        excel_generator._create_workbook()
        assert excel_generator.workbook is not None
        assert excel_generator.ws is not None
        assert excel_generator.ws.title == "Reporte"
    
    def test_create_header(self, excel_generator, user):
        """Test _create_header method."""
        excel_generator._create_workbook()
        excel_generator._create_header("Test Title", user)
        
        assert excel_generator.ws['A1'].value == "Test Title"
        assert excel_generator.ws['A1'].font.bold is True
        assert excel_generator.ws['A3'].value is not None
        assert excel_generator.ws['A4'].value is not None
    
    def test_create_header_without_user(self, excel_generator):
        """Test _create_header without user."""
        excel_generator._create_workbook()
        excel_generator._create_header("Test Title", None)
        
        assert excel_generator.ws['A1'].value == "Test Title"
        assert excel_generator.ws['A3'].value is not None
        assert excel_generator.ws['A4'].value is None or excel_generator.ws['A4'].value == ""
    
    def test_apply_header_style(self, excel_generator):
        """Test _apply_header_style method."""
        excel_generator._create_workbook()
        excel_generator._apply_header_style(1, 5)
        
        for col in range(1, 6):
            cell = excel_generator.ws.cell(row=1, column=col)
            assert cell.fill is not None
            assert cell.font.bold is True
            assert cell.border is not None
    
    def test_apply_cell_border(self, excel_generator):
        """Test _apply_cell_border method."""
        excel_generator._create_workbook()
        excel_generator._apply_cell_border(1, 1)
        
        cell = excel_generator.ws.cell(row=1, column=1)
        assert cell.border is not None
    
    def test_adjust_column_widths(self, excel_generator):
        """Test _adjust_column_widths method."""
        excel_generator._create_workbook()
        excel_generator.ws['A1'] = 'Test'
        excel_generator.ws['B1'] = 'Test'
        
        column_widths = {'A': 20, 'B': 30}
        excel_generator._adjust_column_widths(column_widths)
        
        assert excel_generator.ws.column_dimensions['A'].width == 20
        assert excel_generator.ws.column_dimensions['B'].width == 30
    
    def test_adjust_column_widths_with_error(self, excel_generator):
        """Test _adjust_column_widths with invalid column."""
        excel_generator._create_workbook()
        column_widths = {'Z': 20}
        excel_generator._adjust_column_widths(column_widths)
        # Should not raise, just log warning
    
    def test_save_to_buffer(self, excel_generator):
        """Test _save_to_buffer method."""
        excel_generator._create_workbook()
        excel_generator.ws['A1'] = 'Test'
        
        content = excel_generator._save_to_buffer()
        assert isinstance(content, bytes)
        assert len(content) > 0
    
    def test_save_to_buffer_empty_workbook(self, excel_generator):
        """Test _save_to_buffer with empty workbook."""
        excel_generator._create_workbook()
        content = excel_generator._save_to_buffer()
        assert isinstance(content, bytes)
        assert len(content) > 0
    
    def test_save_to_buffer_error(self, excel_generator):
        """Test _save_to_buffer with error."""
        excel_generator._create_workbook()
        excel_generator.workbook = None
        
        with pytest.raises(Exception):
            excel_generator._save_to_buffer()
    
    def test_get_client_ip(self, excel_generator):
        """Test _get_client_ip method."""
        request = Mock()
        request.META = {'REMOTE_ADDR': '127.0.0.1'}
        
        ip = excel_generator._get_client_ip(request)
        assert ip == '127.0.0.1'
    
    def test_get_client_ip_x_forwarded_for(self, excel_generator):
        """Test _get_client_ip with X-Forwarded-For."""
        request = Mock()
        request.META = {'HTTP_X_FORWARDED_FOR': '192.168.1.1, 127.0.0.1'}
        
        ip = excel_generator._get_client_ip(request)
        assert ip == '192.168.1.1'
    
    def test_get_client_ip_none(self, excel_generator):
        """Test _get_client_ip with None request."""
        ip = excel_generator._get_client_ip(None)
        assert ip is None
    
    def test_create_section_title(self, excel_generator):
        """Test _create_section_title method."""
        excel_generator._create_workbook()
        excel_generator._create_section_title('A5', "Test Section")
        
        assert excel_generator.ws['A5'].value == "Test Section"
        assert excel_generator.ws['A5'].font.bold is True
    
    def test_create_table_with_data(self, excel_generator):
        """Test _create_table_with_data method."""
        excel_generator._create_workbook()
        data = [
            ['Header1', 'Header2'],
            ['Value1', 'Value2'],
            ['Value3', 'Value4']
        ]
        
        excel_generator._create_table_with_data(
            data,
            start_row=1,
            header_row=1,
            column_widths={'A': 20, 'B': 30}
        )
        
        assert excel_generator.ws['A1'].value == 'Header1'
        assert excel_generator.ws['B1'].value == 'Header2'
        assert excel_generator.ws['A2'].value == 'Value1'
        assert excel_generator.ws['B2'].value == 'Value2'
    
    def test_create_table_with_headers(self, excel_generator):
        """Test _create_table_with_headers method."""
        excel_generator._create_workbook()
        headers = ['Header1', 'Header2']
        data_rows = [['Value1', 'Value2'], ['Value3', 'Value4']]
        
        excel_generator._create_table_with_headers(
            headers,
            data_rows,
            start_row=1,
            column_widths=[20, 30]
        )
        
        assert excel_generator.ws['A1'].value == 'Header1'
        assert excel_generator.ws['B1'].value == 'Header2'
        assert excel_generator.ws['A2'].value == 'Value1'
        assert excel_generator.ws['B2'].value == 'Value2'
    
    def test_apply_date_filters(self, excel_generator):
        """Test _apply_date_filters method."""
        from django.utils import timezone
        from datetime import timedelta
        
        queryset = Mock()
        queryset.filter = Mock(return_value=queryset)
        
        filtros = {
            'fecha_desde': timezone.now().date() - timedelta(days=7),
            'fecha_hasta': timezone.now().date()
        }
        
        result = excel_generator._apply_date_filters(queryset, filtros, 'timestamp')
        assert result is not None
        assert queryset.filter.call_count == 2
    
    def test_apply_date_filters_no_filters(self, excel_generator):
        """Test _apply_date_filters with no filters."""
        queryset = Mock()
        result = excel_generator._apply_date_filters(queryset, None, 'timestamp')
        assert result == queryset
    
    def test_apply_date_filters_only_desde(self, excel_generator):
        """Test _apply_date_filters with only fecha_desde."""
        from django.utils import timezone
        from datetime import timedelta
        
        queryset = Mock()
        queryset.filter = Mock(return_value=queryset)
        
        filtros = {
            'fecha_desde': timezone.now().date() - timedelta(days=7)
        }
        
        result = excel_generator._apply_date_filters(queryset, filtros, 'timestamp')
        assert result is not None
        assert queryset.filter.call_count == 1
    
    def test_create_sheet_with_title(self, excel_generator):
        """Test _create_sheet_with_title method."""
        excel_generator._create_workbook()
        original_ws = excel_generator.ws
        
        new_ws, returned_original = excel_generator._create_sheet_with_title(
            "New Sheet",
            "New Title"
        )
        
        assert new_ws is not None
        assert returned_original == original_ws
        assert new_ws['A1'].value == "New Title"
        assert new_ws['A1'].font.bold is True

