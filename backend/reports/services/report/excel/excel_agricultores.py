"""
Excel generator for farmers report.
Generates Excel reports with farmer and farm information.
"""
import logging
from typing import Optional
from openpyxl.styles import Font, Alignment, PatternFill

from .excel_base import ExcelBaseGenerator

logger = logging.getLogger("cacaoscan.services.report.excel.agricultores")


class ExcelAgricultoresGenerator(ExcelBaseGenerator):
    """
    Generator for farmers Excel reports.
    """
    
    def _setup_headers(self):
        """Setup headers and apply styling."""
        columns = [
            'Agricultor', 'Email', 'Teléfono', 'Departamento', 'Municipio',
            'Finca', 'Hectáreas', 'Estado Finca', 'Fecha Registro Finca'
        ]
        self.ws.append(columns)
        
        try:
            header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in self.ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
        except Exception as style_error:
            logger.warning(f"Error applying header styles (non-critical): {style_error}")
    
    def _get_farmers_list(self):
        """Get list of farmers with error handling."""
        from django.contrib.auth.models import User
        
        farmers_query = User.objects.filter(is_superuser=False, is_staff=False)
        
        try:
            farmers = farmers_query.prefetch_related('fincas_app_fincas')
            farmers_list = list(farmers)
            logger.info(f"[INFO] Obtenidos {len(farmers_list)} agricultores con prefetch")
            return farmers_list
        except Exception as prefetch_error:
            logger.warning(f"Error with prefetch_related, trying without prefetch: {prefetch_error}")
            try:
                farmers_list = list(farmers_query)
                logger.info(f"Obtenidos {len(farmers_list)} agricultores sin prefetch")
                return farmers_list
            except Exception as query_error:
                logger.error(f"Error getting users: {query_error}")
                return []
    
    def _get_farmer_name(self, farmer):
        """Get farmer name safely."""
        try:
            if farmer.first_name or farmer.last_name:
                name = f"{farmer.first_name or ''} {farmer.last_name or ''}".strip()
            if not name:
                name = str(farmer.username) if farmer.username else f'Usuario {farmer.id}'
            return name
        except Exception:
            return f'Usuario {farmer.id}'
    
    def _get_farmer_phone(self, farmer):
        """Get farmer phone safely."""
        try:
            if hasattr(farmer, 'auth_profile'):
                profile = getattr(farmer, 'auth_profile', None)
                if profile and hasattr(profile, 'phone_number'):
                    return str(profile.phone_number) if profile.phone_number else ''
        except Exception:
            pass
        return ''
    
    def _get_farmer_farms(self, farmer):
        """Get farmer farms safely."""
        try:
            if hasattr(farmer, 'fincas_app_fincas'):
                fincas_queryset = farmer.fincas_app_fincas.all()
                return list(fincas_queryset)
        except Exception as finca_error:
            logger.warning(f"Error getting farms for user {farmer.id}: {finca_error}")
        return []
    
    def _safe_str_attr(self, obj, attr, default=''):
        """Safely get string attribute."""
        try:
            value = getattr(obj, attr, None)
            return str(value) if value else default
        except Exception:
            return default
    
    def _safe_float_attr(self, obj, attr, default=0.0):
        """Safely get float attribute."""
        try:
            value = getattr(obj, attr, None)
            return float(value) if value is not None else default
        except (ValueError, TypeError, AttributeError):
            return default
    
    def _safe_date_str(self, obj, attr, date_format='%Y-%m-%d', default=''):
        """Safely get date string."""
        try:
            if hasattr(obj, attr):
                date_value = getattr(obj, attr)
                return date_value.strftime(date_format) if date_value else default
        except Exception:
            pass
        return default
    
    def _extract_finca_data(self, finca):
        """Extract farm data safely."""
        return {
            'depto': self._safe_str_attr(finca, 'departamento'),
            'municipio': self._safe_str_attr(finca, 'municipio'),
            'nombre': self._safe_str_attr(finca, 'nombre', 'Sin nombre'),
            'hectareas': self._safe_float_attr(finca, 'hectareas'),
            'estado': "Activa" if getattr(finca, 'activa', False) else "Inactiva",
            'fecha_registro': self._safe_date_str(finca, 'fecha_registro')
        }
    
    def _add_farm_row(self, name, email, phone, finca_data):
        """Add row for farmer with farm data."""
        self.ws.append([
            name,
            email,
            phone,
            finca_data['depto'],
            finca_data['municipio'],
            finca_data['nombre'],
            finca_data['hectareas'],
            finca_data['estado'],
            finca_data['fecha_registro']
        ])
    
    def _add_farmer_without_farms_row(self, name, email, phone, farmer):
        """Add row for farmer without farms."""
        fecha_registro_str = ''
        try:
            if farmer.date_joined:
                fecha_registro_str = farmer.date_joined.strftime('%Y-%m-%d')
        except Exception:
            pass
        
        self.ws.append([
            name,
            email,
            phone,
            '', '', '', '', '',
            fecha_registro_str
        ])
    
    def _process_farmer(self, farmer):
        """Process single farmer and add rows."""
        name = self._get_farmer_name(farmer)
        email = str(farmer.email) if farmer.email else ''
        phone = self._get_farmer_phone(farmer)
        fincas_list = self._get_farmer_farms(farmer)
        
        rows_added = 0
        
        if fincas_list:
            for finca in fincas_list:
                try:
                    finca_data = self._extract_finca_data(finca)
                    self._add_farm_row(name, email, phone, finca_data)
                    rows_added += 1
                except Exception as finca_row_error:
                    logger.warning(f"Error processing farm {getattr(finca, 'id', 'unknown')}: {finca_row_error}")
        else:
            self._add_farmer_without_farms_row(name, email, phone, farmer)
            rows_added += 1
        
        return rows_added
    
    def _finalize_report(self):
        """Finalize report with column widths and alignment."""
        try:
            column_widths = {
                'A': 25, 'B': 30, 'C': 15, 'D': 20, 'E': 20,
                'F': 20, 'G': 12, 'H': 15, 'I': 18,
            }
            self._adjust_column_widths(column_widths)
        except Exception as col_error:
            logger.warning(f"Error adjusting column widths (non-critical): {col_error}")
        
        try:
            for row in self.ws.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        except Exception as align_error:
            logger.warning(f"Error centering headers (non-critical): {align_error}")
    
    def generate_farmers_report(self) -> bytes:
        """
        Generates Excel report of farmers with their farms.
        
        Returns:
            bytes: Excel file content
        """
        try:
            logger.info("[INFO] Iniciando generación de reporte Excel de agricultores")
            
            self._create_workbook("Agricultores")
            self._setup_headers()
            
            farmers_list = self._get_farmers_list()
            
            rows_added = 0
            for farmer in farmers_list:
                try:
                    rows_added += self._process_farmer(farmer)
                except Exception as farmer_error:
                    logger.warning(f"Error processing farmer {getattr(farmer, 'id', 'unknown')}: {farmer_error}")
                    continue
            
            logger.info(f"[INFO] Procesados {rows_added} filas de datos")
            
            self._finalize_report()
            
            content = self._save_to_buffer()
            logger.info(f"[INFO] Reporte Excel de agricultores generado exitosamente - {len(content)} bytes, {rows_added} filas")
            return content
            
        except Exception as e:
            logger.error(f"[ERROR] Error generando reporte Excel de agricultores: {e}", exc_info=True)
            raise

