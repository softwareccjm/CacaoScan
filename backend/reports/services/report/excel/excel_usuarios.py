"""
Excel generator for users report.
Generates Excel reports with user and farm information.
"""
import logging
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from .excel_base import ExcelBaseGenerator

logger = logging.getLogger("cacaoscan.services.report.excel.usuarios")


class ExcelUsuariosGenerator(ExcelBaseGenerator):
    """
    Generator for users Excel reports.
    """
    
    def _get_styles(self):
        """Get reusable Excel styles."""
        return {
            'bold_font': Font(bold=True, color="FFFFFF"),
            'header_fill': PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid"),
            'thin_border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            ),
            'align_center': Alignment(horizontal='center', vertical='center'),
            'align_vertical': Alignment(vertical='center')
        }
    
    def _setup_headers(self, styles):
        """Setup and style headers."""
        headers = [
            'ID Usuario', 'Nombre', 'Correo', 'Rol', 'Activo', 'Fecha Registro',
            'Finca', 'Departamento', 'Municipio', 'Área (ha)', 'Latitud', 'Longitud'
        ]
        self.ws.append(headers)
        
        for col in self.ws[1]:
            col.font = styles['bold_font']
            col.fill = styles['header_fill']
            col.alignment = styles['align_center']
            col.border = styles['thin_border']
        
        return headers
    
    def _get_user_role(self, user):
        """Determine user role."""
        if user.is_superuser or user.is_staff:
            return 'admin'
        if user.groups.filter(name='analyst').exists():
            return 'analyst'
        return 'farmer'
    
    def _add_user_row(self, user, rol, finca=None):
        """Add a row for user with optional farm data."""
        base_row = [
            user.id,
            f"{user.first_name} {user.last_name}".strip() or user.username,
            user.email,
            rol,
            'Sí' if user.is_active else 'No',
            user.date_joined.strftime('%Y-%m-%d'),
        ]
        
        if finca:
            base_row.extend([
                finca.nombre,
                finca.departamento or '',
                finca.municipio or '',
                float(finca.hectareas) if finca.hectareas else '',
                float(finca.coordenadas_lat) if finca.coordenadas_lat is not None else '',
                float(finca.coordenadas_lng) if finca.coordenadas_lng is not None else '',
            ])
        else:
            base_row.extend(['Sin fincas', '', '', '', '', ''])
        
        self.ws.append(base_row)
    
    def _add_user_data(self, users):
        """Add user data rows to worksheet."""
        for user in users:
            fincas = user.fincas_app_fincas.all()
            rol = self._get_user_role(user)
            
            if fincas.exists():
                for finca in fincas:
                    self._add_user_row(user, rol, finca)
            else:
                self._add_user_row(user, rol)
    
    def _apply_data_formatting(self, styles, headers):
        """Apply borders and alignment to data cells."""
        for row in self.ws.iter_rows(min_row=2, max_row=self.ws.max_row, max_col=len(headers)):
            for cell in row:
                cell.border = styles['thin_border']
                cell.alignment = styles['align_vertical']
    
    def _auto_adjust_columns(self):
        """Auto-adjust column widths."""
        for col in self.ws.columns:
            max_length = 0
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            self.ws.column_dimensions[col[0].column_letter].width = adjusted_width
    
    def _show_empty_message(self, styles):
        """Show message when no users are available."""
        self.ws.merge_cells('A1:L2')
        cell = self.ws['A1']
        cell.value = "Sin registros disponibles"
        cell.font = Font(bold=True, size=14, color="FF0000")
        cell.alignment = styles['align_center']
        cell.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
        self.ws.row_dimensions[1].height = 40
    
    def generate_users_report(self) -> bytes:
        """
        Generates professional Excel report of users with their associated farms.
        Includes professional visual formatting (colors, borders, alignment).
        If no users, shows "Sin registros disponibles" message.
        
        Returns:
            bytes: Excel file content
        """
        try:
            from django.contrib.auth.models import User
            
            self._create_workbook("Usuarios y Fincas")
            styles = self._get_styles()
            headers = self._setup_headers(styles)
            
            users = User.objects.all().order_by('-date_joined').select_related(
                'auth_profile', 'auth_email_token'
            ).prefetch_related('fincas_app_fincas', 'groups')
            
            if users.exists():
                self._add_user_data(users)
                self._apply_data_formatting(styles, headers)
                self._auto_adjust_columns()
            else:
                self._show_empty_message(styles)
            
            content = self._save_to_buffer()
            
            if not content or len(content) < 100:
                raise ValueError("Generated Excel file is empty or corrupt")
            
            logger.info(f"Reporte Excel de usuarios y fincas generado correctamente ({len(content)} bytes)")
            return content
            
        except Exception as e:
            logger.error(f"Error generando reporte Excel de usuarios: {e}", exc_info=True)
            raise

