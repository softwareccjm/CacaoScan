"""
Base Excel generator for CacaoScan.
Contains common functionality and styles for Excel report generation.
"""
import logging
import io
from datetime import datetime
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

logger = logging.getLogger("cacaoscan.services.report.excel.base")

# Excel column constants (shared across all Excel generators)
EXCEL_COL_METRIC = 'Métrica'
EXCEL_COL_VALUE = 'Valor'
EXCEL_TOTAL_ANALISIS = 'Total de Análisis'
EXCEL_AVG_CONFIDENCE = 'Confianza Promedio'
EXCEL_AVG_ALTO = 'Alto Promedio'
EXCEL_AVG_ANCHO = 'Ancho Promedio'
EXCEL_AVG_GROSOR = 'Grosor Promedio'
EXCEL_AVG_PESO = 'Peso Promedio'
EXCEL_TOTAL_LOTES = 'Total de Lotes'
EXCEL_LOTES_ACTIVOS = 'Lotes Activos'
DATE_TIME_FORMAT = '%d/%m/%Y %H:%M'


class ExcelBaseGenerator:
    """
    Base class for Excel report generation.
    Provides common functionality and styles.
    """
    
    # Common style constants to avoid duplication
    HEADER_COLOR = "2F4F4F"
    HEADER_TEXT_COLOR = "FFFFFF"
    TITLE_COLOR = "2F4F4F"
    
    def __init__(self):
        self.workbook = None
        self.ws = None
    
    def _get_thin_border(self) -> Border:
        """
        Returns a standard thin border style.
        
        Returns:
            Border: Thin border object
        """
        return Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def _get_header_fill(self) -> PatternFill:
        """
        Returns standard header fill style.
        
        Returns:
            PatternFill: Header fill object
        """
        return PatternFill(
            start_color=self.HEADER_COLOR,
            end_color=self.HEADER_COLOR,
            fill_type="solid"
        )
    
    def _get_header_font(self) -> Font:
        """
        Returns standard header font style.
        
        Returns:
            Font: Header font object
        """
        return Font(bold=True, color=self.HEADER_TEXT_COLOR)
    
    def _get_alignment_map(self) -> dict:
        """
        Returns alignment mapping dictionary.
        
        Returns:
            dict: Mapping of alignment strings to Alignment objects
        """
        return {
            'center': Alignment(horizontal='center'),
            'left': Alignment(horizontal='left'),
            'right': Alignment(horizontal='right')
        }
    
    def _create_workbook(self, sheet_title: str = "Reporte"):
        """Creates a new workbook and worksheet."""
        self.workbook = Workbook()
        self.ws = self.workbook.active
        self.ws.title = sheet_title
    
    def _create_header(self, title: str, user=None):
        """
        Creates report header.
        
        Args:
            title: Report title
            user: User who requested the report
        """
        # Main title
        self.ws['A1'] = title
        self.ws['A1'].font = Font(size=16, bold=True, color=self.TITLE_COLOR)
        self.ws['A1'].alignment = Alignment(horizontal='center')
        self.ws.merge_cells('A1:F1')
        
        # Report information
        self.ws['A3'] = f"Generado el: {timezone.now().strftime('%d/%m/%Y %H:%M')}"
        self.ws['A3'].font = Font(size=10, italic=True)
        
        if user:
            self.ws['A4'] = f"Usuario: {user.get_full_name() or user.username}"
            self.ws['A4'].font = Font(size=10, italic=True)
        
        # Space
        self.ws['A6'] = ""
    
    def _apply_header_style(self, row_num: int, num_cols: int):
        """
        Applies header style to a row.
        
        Args:
            row_num: Row number
            num_cols: Number of columns
        """
        header_fill = self._get_header_fill()
        header_font = self._get_header_font()
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = self._get_thin_border()
        
        for col in range(1, num_cols + 1):
            cell = self.ws.cell(row=row_num, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
    
    def _apply_cell_border(self, row_num: int, col_num: int):
        """
        Applies border to a cell.
        
        Args:
            row_num: Row number
            col_num: Column number
        """
        thin_border = self._get_thin_border()
        cell = self.ws.cell(row=row_num, column=col_num)
        cell.border = thin_border
    
    def _adjust_column_widths(self, column_widths: dict):
        """
        Adjusts column widths.
        
        Args:
            column_widths: Dict with column letters as keys and widths as values
        """
        for col, width in column_widths.items():
            try:
                self.ws.column_dimensions[col].width = width
            except Exception as e:
                logger.warning(f"Error adjusting column width {col}: {e}")
    
    def _save_to_buffer(self) -> bytes:
        """
        Saves workbook to buffer and returns content.
        
        Returns:
            bytes: Excel file content
        """
        buffer = io.BytesIO()
        try:
            self.workbook.save(buffer)
        except Exception as save_error:
            logger.error(f"Error saving workbook: {save_error}", exc_info=True)
            raise
        
        buffer.seek(0)
        content = buffer.getvalue()
        
        # Validate content
        if not content:
            logger.error("Generated Excel file is empty")
            raise ValueError("Generated Excel file is empty")
        
        if len(content) < 50:
            logger.warning(f"Generated Excel file is very small: {len(content)} bytes")
        
        return content
    
    def _get_client_ip(self, request):
        """Gets client IP address."""
        if not request:
            return None
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0]
        else:
            ip = request.META.get('REMOTE_ADDR')
        return ip
    
    def _create_section_title(self, cell_address: str, title: str):
        """
        Creates a section title.
        
        Args:
            cell_address: Cell address (e.g., 'A8')
            title: Section title text
        """
        self.ws[cell_address] = title
        self.ws[cell_address].font = Font(size=14, bold=True, color=self.TITLE_COLOR)
    
    def _create_table_with_data(
        self,
        data: list[list],
        start_row: int,
        header_row: int,
        column_widths: dict = None,
        header_alignment: str = 'center',
        body_alignment: str = 'center'
    ):
        """
        Creates a table with headers and data rows.
        
        Args:
            data: List of rows, first row should be headers
            start_row: Starting row number
            header_row: Row number for headers (usually same as start_row)
            column_widths: Dict with column letters as keys and widths as values
            header_alignment: Alignment for header cells ('center', 'left', 'right')
            body_alignment: Alignment for body cells ('center', 'left', 'right')
        """
        thin_border = self._get_thin_border()
        header_fill = self._get_header_fill()
        header_font = self._get_header_font()
        
        alignment_map = self._get_alignment_map()
        header_align = alignment_map.get(header_alignment, Alignment(horizontal='center'))
        body_align = alignment_map.get(body_alignment, Alignment(horizontal='center'))
        
        for row_num, row_data in enumerate(data, start_row):
            is_header = row_num == header_row
            for col_num, cell_value in enumerate(row_data, 1):
                cell = self.ws.cell(row=row_num, column=col_num, value=cell_value)
                
                if is_header:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_align
                else:
                    cell.alignment = body_align
                
                cell.border = thin_border
        
        if column_widths:
            self._adjust_column_widths(column_widths)
    
    def _create_table_with_headers(
        self,
        headers: list,
        data_rows: list[list],
        start_row: int,
        column_widths: list = None,
        header_alignment: str = 'center',
        body_alignment: str = 'center'
    ):
        """
        Creates a table with separate headers and data rows.
        
        Args:
            headers: List of header strings
            data_rows: List of data rows
            start_row: Starting row number for headers
            column_widths: List of widths (will be mapped to column letters)
            header_alignment: Alignment for header cells
            body_alignment: Alignment for body cells
        """
        thin_border = self._get_thin_border()
        header_fill = self._get_header_fill()
        header_font = self._get_header_font()
        
        alignment_map = self._get_alignment_map()
        header_align = alignment_map.get(header_alignment, Alignment(horizontal='center'))
        body_align = alignment_map.get(body_alignment, Alignment(horizontal='center'))
        
        # Create headers
        for col_num, header in enumerate(headers, 1):
            cell = self.ws.cell(row=start_row, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        
        # Create data rows
        for row_offset, row_data in enumerate(data_rows, 1):
            row_num = start_row + row_offset
            for col_num, value in enumerate(row_data, 1):
                cell = self.ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = body_align
                cell.border = thin_border
        
        # Adjust column widths
        if column_widths:
            width_dict = {}
            for i, width in enumerate(column_widths, 1):
                col_letter = chr(64 + i)
                width_dict[col_letter] = width
            self._adjust_column_widths(width_dict)
    
    def _apply_date_filters(self, queryset, filtros, date_field_name: str):
        """
        Applies date filters to a queryset.
        
        Args:
            queryset: Django queryset to filter
            filtros: Filters dictionary
            date_field_name: Name of the date field in the model (e.g., 'timestamp', 'created_at', 'login_time')
            
        Returns:
            Filtered queryset
        """
        if not filtros:
            return queryset
        
        if filtros.get('fecha_desde'):
            queryset = queryset.filter(**{f"{date_field_name}__date__gte": filtros['fecha_desde']})
        if filtros.get('fecha_hasta'):
            queryset = queryset.filter(**{f"{date_field_name}__date__lte": filtros['fecha_hasta']})
        
        return queryset
    
    def _create_sheet_with_title(self, sheet_name: str, title: str, merge_range: str = 'A1:F1'):
        """
        Creates a new sheet with a formatted title.
        
        Args:
            sheet_name: Name of the new sheet
            title: Title text to display
            merge_range: Cell range to merge for title (default: 'A1:F1')
            
        Returns:
            Tuple of (new_worksheet, original_worksheet)
        """
        new_ws = self.workbook.create_sheet(sheet_name)
        original_ws = self.ws
        self.ws = new_ws
        
        # Title
        new_ws['A1'] = title
        new_ws['A1'].font = Font(size=16, bold=True, color=self.TITLE_COLOR)
        new_ws['A1'].alignment = Alignment(horizontal='center')
        new_ws.merge_cells(merge_range)
        
        return new_ws, original_ws

