"""
Excel generator for analysis reports.
Generates Excel reports for quality, farm, audit, and custom analysis.
"""
import logging
import io
from django.utils import timezone
from django.db.models import Count, Avg
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, Reference

from .excel_base import ExcelBaseGenerator
from ..report_stats import (
    apply_prediction_filters,
    get_quality_stats,
    get_lotes_stats,
    get_activity_stats,
    get_login_stats
)

# Import models safely
from api.utils.model_imports import get_models_safely

models = get_models_safely({
    'CacaoPrediction': 'images_app.models.CacaoPrediction',
    'Finca': 'fincas_app.models.Finca',
    'Lote': 'fincas_app.models.Lote',
    'ActivityLog': 'audit.models.ActivityLog'
})
CacaoPrediction = models['CacaoPrediction']
Finca = models['Finca']
Lote = models['Lote']
ActivityLog = models['ActivityLog']

from audit.models import LoginHistory

logger = logging.getLogger("cacaoscan.services.report.excel.analisis")

# Import shared constants from base
from .excel_base import (
    EXCEL_COL_METRIC,
    EXCEL_COL_VALUE,
    EXCEL_TOTAL_ANALISIS,
    EXCEL_AVG_CONFIDENCE,
    EXCEL_AVG_ALTO,
    EXCEL_AVG_ANCHO,
    EXCEL_AVG_GROSOR,
    EXCEL_AVG_PESO,
    EXCEL_TOTAL_LOTES,
    EXCEL_LOTES_ACTIVOS,
    DATE_TIME_FORMAT
)

# Sheet and chart titles
EXCEL_QUALITY_DISTRIBUTION_TITLE = "Distribución de Calidad"


class ExcelAnalisisGenerator(ExcelBaseGenerator):
    """
    Generator for analysis Excel reports (quality, farm, audit, custom).
    """
    
    def generate_quality_report(self, user, filtros=None) -> bytes:
        """
        Generates quality report in Excel.
        
        Args:
            user: User requesting the report
            filtros: Filters to apply
            
        Returns:
            bytes: Excel file content
        """
        try:
            self._create_workbook("Reporte de Calidad")
            
            # Apply filters
            queryset = apply_prediction_filters(CacaoPrediction.objects.all(), filtros)
            
            # Create header
            self._create_header("Reporte de Calidad de Granos de Cacao", user)
            
            # General statistics
            stats = get_quality_stats(queryset)
            self._create_stats_section(stats)
            
            # Detailed analyses table
            self._create_detailed_analyses_table(queryset)
            
            # Quality distribution chart
            self._create_quality_chart(stats)
            
            # Summary sheet
            self._create_summary_sheet(stats, user)
            
            # Save to buffer
            return self._save_to_buffer()
            
        except Exception as e:
            logger.error(f"Error generando reporte Excel de calidad: {e}")
            raise
    
    def generate_finca_report(self, finca_id, user, filtros=None) -> bytes:
        """
        Generates farm report in Excel.
        
        Args:
            finca_id: ID de la finca
            user: Usuario que solicita el reporte
            filtros: Filtros a aplicar (reservado para uso futuro)
        
        Args:
            finca_id: Farm ID
            user: User requesting the report
            filtros: Filters to apply
            
        Returns:
            bytes: Excel file content
        """
        # Suppress unused parameter warnings - reserved for future use
        _ = user
        _ = filtros
        
        try:
            finca = Finca.objects.get(id=finca_id)
            self._create_workbook(f"Finca {finca.nombre}")
            
            # Create header
            self._create_header(f"Reporte de Finca: {finca.nombre}", user)
            
            # Farm information
            self._create_finca_info_section(finca)
            
            # Lot statistics
            lotes_stats = get_lotes_stats(finca)
            self._create_lotes_stats_section(lotes_stats)
            
            # Analysis by lot
            self._create_lotes_analysis_section(finca)
            
            # Detailed lots sheet
            self._create_detailed_lotes_sheet(finca)
            
            # Save to buffer
            return self._save_to_buffer()
            
        except Exception as e:
            logger.error(f"Error generando reporte Excel de finca: {e}")
            raise
    
    def generate_audit_report(self, user, filtros=None) -> bytes:
        """
        Generates audit report in Excel.
        
        Args:
            user: User requesting the report
            filtros: Filters to apply
            
        Returns:
            bytes: Excel file content
        """
        try:
            self._create_workbook("Auditoría")
            
            # Create header
            self._create_header("Reporte de Auditoría del Sistema", user)
            
            # Activity statistics
            activity_stats = get_activity_stats(filtros)
            self._create_activity_stats_section(activity_stats)
            
            # Login statistics
            login_stats = get_login_stats(filtros)
            self._create_login_stats_section(login_stats)
            
            # Detailed activities sheet
            self._create_detailed_activities_sheet(filtros)
            
            # Detailed logins sheet
            self._create_detailed_logins_sheet(filtros)
            
            # Save to buffer
            return self._save_to_buffer()
            
        except Exception as e:
            logger.error(f"Error generando reporte Excel de auditoría: {e}")
            raise
    
    def generate_custom_report(self, user, tipo_reporte, parametros, filtros=None) -> bytes:
        """
        Generates custom report in Excel.
        
        Args:
            user: User requesting the report
            tipo_reporte: Report type
            parametros: Report parameters
            filtros: Filters to apply
            
        Returns:
            bytes: Excel file content
        """
        try:
            self._create_workbook("Reporte Personalizado")
            
            # Create header
            self._create_header(f"Reporte Personalizado: {tipo_reporte}", user)
            
            # Generate according to type
            if tipo_reporte == 'calidad':
                queryset = apply_prediction_filters(CacaoPrediction.objects.all(), filtros)
                stats = get_quality_stats(queryset)
                self._create_custom_quality_section(stats, parametros)
            elif tipo_reporte == 'finca':
                if parametros.get('finca_id'):
                    finca = Finca.objects.get(id=parametros['finca_id'])
                    self._create_custom_finca_section(finca, parametros)
            elif tipo_reporte == 'auditoria':
                activity_stats = get_activity_stats(filtros)
                self._create_custom_audit_section(activity_stats, parametros)
            
            # Save to buffer
            return self._save_to_buffer()
            
        except Exception as e:
            logger.error(f"Error generando reporte Excel personalizado: {e}")
            raise
    
    
    def _create_stats_section(self, stats):
        """Create statistics section."""
        self._create_section_title('A8', "Estadísticas Generales")
        
        data = [
            [EXCEL_COL_METRIC, EXCEL_COL_VALUE],
            [EXCEL_TOTAL_ANALISIS, stats['total_analyses']],
            [EXCEL_AVG_CONFIDENCE, f"{stats['avg_confidence']}%"],
            [EXCEL_AVG_ALTO, f"{stats['avg_dimensions']['alto']} mm"],
            [EXCEL_AVG_ANCHO, f"{stats['avg_dimensions']['ancho']} mm"],
            [EXCEL_AVG_GROSOR, f"{stats['avg_dimensions']['grosor']} mm"],
            [EXCEL_AVG_PESO, f"{stats['avg_weight']} g"],
        ]
        
        self._create_table_with_data(
            data,
            start_row=10,
            header_row=10,
            column_widths={'A': 20, 'B': 15}
        )
    
    def _create_detailed_analyses_table(self, queryset):
        """Create detailed analyses table."""
        self._create_section_title('A18', "Análisis Detallados")
        
        headers = ['ID', 'Usuario', 'Finca', 'Región', 'Fecha', 'Alto (mm)', 'Ancho (mm)', 'Grosor (mm)', 'Peso (g)', 'Confianza']
        analyses = queryset.select_related('image', 'image__user').order_by('-created_at')[:100]
        
        data_rows = [
            [
                analysis.id,
                analysis.image.user.username,
                analysis.image.finca or 'N/A',
                analysis.image.region or 'N/A',
                analysis.created_at.strftime(DATE_TIME_FORMAT),
                round(float(analysis.alto_mm), 2),
                round(float(analysis.ancho_mm), 2),
                round(float(analysis.grosor_mm), 2),
                round(float(analysis.peso_g), 2),
                f"{analysis.average_confidence:.2%}"
            ]
            for analysis in analyses
        ]
        
        self._create_table_with_headers(
            headers,
            data_rows,
            start_row=20,
            column_widths=[8, 12, 15, 12, 18, 10, 10, 10, 10, 12]
        )
    
    def _create_quality_chart(self, stats):
        """Create quality distribution chart."""
        if not stats.get('quality_distribution'):
            return
        
        chart_ws, original_ws = self._create_sheet_with_title(
            EXCEL_QUALITY_DISTRIBUTION_TITLE,
            EXCEL_QUALITY_DISTRIBUTION_TITLE
        )
        
        headers = ["Categoría", "Cantidad", "Porcentaje"]
        total = stats.get('total_analyses', sum(stats['quality_distribution'].values()))
        data_rows = [
            [
                category,
                count,
                f"{(count / total * 100) if total > 0 else 0:.1f}%"
            ]
            for category, count in stats['quality_distribution'].items()
        ]
        
        self._create_table_with_headers(
            headers,
            data_rows,
            start_row=3,
            column_widths=[20, 12, 12]
        )
        
        # Create bar chart
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = EXCEL_QUALITY_DISTRIBUTION_TITLE
        chart.y_axis.title = 'Cantidad'
        chart.x_axis.title = 'Categoría'
        
        last_row = len(data_rows) + 2  # +2 because headers are at row 3, data starts at row 4
        data = Reference(chart_ws, min_col=2, min_row=3, max_row=last_row, max_col=2)
        cats = Reference(chart_ws, min_col=1, min_row=4, max_row=last_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        chart_ws.add_chart(chart, "E2")
        self.ws = original_ws
    
    def _create_summary_sheet(self, stats, user):
        """Create summary sheet."""
        summary_ws, _ = self._create_sheet_with_title(
            "Resumen",
            "Resumen Ejecutivo",
            'A1:D1'
        )
        
        # Report information
        summary_ws['A3'] = f"Generado el: {timezone.now().strftime(DATE_TIME_FORMAT)}"
        summary_ws['A3'].font = Font(size=10, italic=True)
        
        summary_ws['A4'] = f"Usuario: {user.get_full_name() or user.username}"
        summary_ws['A4'].font = Font(size=10, italic=True)
        
        # Key metrics summary
        summary_ws['A6'] = "Métricas Clave"
        summary_ws['A6'].font = Font(size=14, bold=True, color="2F4F4F")
        
        key_metrics = [
            [EXCEL_TOTAL_ANALISIS, stats['total_analyses']],
            [EXCEL_AVG_CONFIDENCE, f"{stats['avg_confidence']}%"],
            [EXCEL_AVG_ALTO, f"{stats['avg_dimensions']['alto']} mm"],
            [EXCEL_AVG_ANCHO, f"{stats['avg_dimensions']['ancho']} mm"],
            [EXCEL_AVG_GROSOR, f"{stats['avg_dimensions']['grosor']} mm"],
            [EXCEL_AVG_PESO, f"{stats['avg_weight']} g"],
        ]
        
        for row_num, (metric, value) in enumerate(key_metrics, 8):
            summary_ws[f'A{row_num}'] = metric
            summary_ws[f'B{row_num}'] = value
            summary_ws[f'A{row_num}'].font = Font(bold=True)
            summary_ws[f'B{row_num}'].alignment = Alignment(horizontal='center')
        
        # Recommendations
        summary_ws['A15'] = "Recomendaciones"
        summary_ws['A15'].font = Font(size=14, bold=True, color="2F4F4F")
        
        recommendations = []
        if stats['avg_confidence'] < 70:
            recommendations.append("- La confianza promedio es baja. Considere mejorar la calidad de las imágenes.")
        if stats['avg_dimensions']['alto'] < 15 or stats['avg_dimensions']['alto'] > 25:
            recommendations.append("- Las dimensiones están fuera del rango óptimo. Revise el proceso de cosecha.")
        if stats['avg_weight'] < 1.0 or stats['avg_weight'] > 2.5:
            recommendations.append("- El peso promedio no está en el rango esperado. Verifique la madurez.")
        
        if not recommendations:
            recommendations.append("- Los indicadores están dentro de rangos aceptables. Mantenga las buenas prácticas.")
        
        for row_num, rec in enumerate(recommendations, 17):
            summary_ws[f'A{row_num}'] = rec
            summary_ws[f'A{row_num}'].font = Font(size=10)
        
        # Adjust column widths
        summary_ws.column_dimensions['A'].width = 25
        summary_ws.column_dimensions['B'].width = 15
    
    def _create_finca_info_section(self, finca):
        """Create farm information section."""
        self._create_section_title('A8', "Información de la Finca")
        
        finca_data = [
            ['Campo', 'Valor'],
            ['Nombre', finca.nombre],
            ['Ubicación', finca.ubicacion_completa],
            ['Hectáreas', f"{finca.hectareas} ha"],
            ['Agricultor', finca.agricultor.get_full_name() or finca.agricultor.username],
            [EXCEL_TOTAL_LOTES, str(finca.total_lotes)],
            [EXCEL_LOTES_ACTIVOS, str(finca.lotes_activos)],
            [EXCEL_TOTAL_ANALISIS, str(finca.total_analisis)],
            ['Calidad Promedio', f"{finca.calidad_promedio}%"],
        ]
        
        self._create_table_with_data(
            finca_data,
            start_row=10,
            header_row=10,
            column_widths={'A': 20, 'B': 30},
            body_alignment='left'
        )
    
    
    def _create_lotes_stats_section(self, stats):
        """Create lot statistics section."""
        self._create_section_title('A20', "Estadísticas de Lotes")
        
        data = [
            [EXCEL_COL_METRIC, EXCEL_COL_VALUE],
            [EXCEL_TOTAL_LOTES, str(stats['total_lotes'])],
            [EXCEL_LOTES_ACTIVOS, str(stats['lotes_activos'])],
            ['Área Total', f"{stats['total_area']:.2f} ha"],
            ['Variedades', str(len(stats['variedades']))],
        ]
        
        self._create_table_with_data(
            data,
            start_row=22,
            header_row=22,
            column_widths={'A': 20, 'B': 15}
        )
    
    def _create_lotes_analysis_section(self, finca):
        """Create analysis by lot section."""
        self._create_section_title('A28', "Análisis por Lote")
        
        lotes = finca.lotes.all()
        
        if lotes.exists():
            headers = ['Lote', 'Variedad', 'Estado', 'Área (ha)', 'Análisis', 'Calidad (%)']
            data_rows = [
                [
                    lote.identificador,
                    lote.variedad,
                    lote.get_estado_display(),
                    f"{lote.area_hectareas:.2f}",
                    str(lote.total_analisis),
                    f"{lote.calidad_promedio:.1f}",
                ]
                for lote in lotes
            ]
            
            self._create_table_with_headers(
                headers,
                data_rows,
                start_row=30,
                column_widths=[12, 15, 12, 12, 10, 12]
            )
    
    def _create_detailed_lotes_sheet(self, finca):
        """Create detailed lots sheet."""
        lotes_ws = self.workbook.create_sheet("Lotes Detallados")
        original_ws = self.ws
        self.ws = lotes_ws
        
        # Title
        lotes_ws['A1'] = f"Análisis Detallados - Finca {finca.nombre}"
        lotes_ws['A1'].font = Font(size=16, bold=True, color="2F4F4F")
        lotes_ws['A1'].alignment = Alignment(horizontal='center')
        lotes_ws.merge_cells('A1:H1')
        
        headers = ['Lote', 'Variedad', 'Estado', 'Área (ha)', 'Fecha Plantación', 'Análisis', 'Calidad (%)', 'Observaciones']
        lotes = finca.lotes.all()
        data_rows = [
            [
                lote.identificador,
                lote.variedad,
                lote.get_estado_display(),
                f"{lote.area_hectareas:.2f}",
                lote.fecha_plantacion.strftime('%d/%m/%Y'),
                str(lote.total_analisis),
                f"{lote.calidad_promedio:.1f}",
                lote.descripcion or 'Sin observaciones',
            ]
            for lote in lotes
        ]
        
        self._create_table_with_headers(
            headers,
            data_rows,
            start_row=3,
            column_widths=[12, 15, 12, 12, 15, 10, 12, 25]
        )
        
        self.ws = original_ws
    
    
    def _create_activity_stats_section(self, stats):
        """Create activity statistics section."""
        self._create_section_title('A8', "Estadísticas de Actividad")
        
        data = [
            [EXCEL_COL_METRIC, EXCEL_COL_VALUE],
            ['Total de Actividades', str(stats['total_activities'])],
            ['Actividades Hoy', str(stats['activities_today'])],
        ]
        
        self._create_table_with_data(
            data,
            start_row=10,
            header_row=10,
            column_widths={'A': 20, 'B': 15}
        )
    
    
    def _create_login_stats_section(self, stats):
        """Create login statistics section."""
        self._create_section_title('A14', "Estadísticas de Logins")
        
        data = [
            [EXCEL_COL_METRIC, EXCEL_COL_VALUE],
            ['Total de Logins', str(stats['total_logins'])],
            ['Logins Exitosos', str(stats['successful_logins'])],
            ['Logins Fallidos', str(stats['failed_logins'])],
            ['Tasa de éxito', f"{stats['success_rate']:.1f}%"],
        ]
        
        self._create_table_with_data(
            data,
            start_row=16,
            header_row=16,
            column_widths={'A': 20, 'B': 15}
        )
    
    def _create_detailed_activities_sheet(self, filtros):
        """Create detailed activities sheet."""
        _, original_ws = self._create_sheet_with_title(
            "Actividades Detalladas",
            "Actividades del Sistema"
        )
        
        queryset = ActivityLog.objects.select_related('user').order_by('-timestamp')[:100]
        queryset = self._apply_date_filters(queryset, filtros, 'timestamp')
        
        headers = ['Fecha', 'Usuario', 'Acción', 'Tipo Recurso', 'Detalles', 'IP']
        data_rows = []
        for activity in queryset:
            if isinstance(activity.details, dict):
                descripcion = activity.details.get('description', '')
            else:
                descripcion = str(activity.details)
            
            descripcion_truncada = descripcion[:50] + '...' if len(descripcion) > 50 else descripcion
            
            data_rows.append([
                activity.timestamp.strftime(DATE_TIME_FORMAT),
                activity.user.username if activity.user else 'Anónimo',
                activity.action,
                activity.resource_type or 'N/A',
                descripcion_truncada,
                activity.ip_address or 'N/A',
            ])
        
        self._create_table_with_headers(
            headers,
            data_rows,
            start_row=3,
            column_widths=[18, 12, 12, 12, 30, 15]
        )
        
        self.ws = original_ws
    
    def _create_detailed_logins_sheet(self, filtros):
        """Create detailed logins sheet."""
        _, original_ws = self._create_sheet_with_title(
            "Logins Detallados",
            "Historial de Logins"
        )
        
        queryset = LoginHistory.objects.select_related('user').order_by('-login_time')[:100]
        queryset = self._apply_date_filters(queryset, filtros, 'login_time')
        
        headers = ['Fecha', 'Usuario', 'IP', 'Éxito', 'Duración', 'Razón Fallo']
        data_rows = [
            [
                login.login_time.strftime(DATE_TIME_FORMAT),
                login.usuario.username,
                login.ip_address,
                'Sí' if login.success else 'No',
                login.session_duration_formatted if hasattr(login, 'session_duration_formatted') else 'N/A',
                login.failure_reason or 'N/A',
            ]
            for login in queryset
        ]
        
        self._create_table_with_headers(
            headers,
            data_rows,
            start_row=3,
            column_widths=[18, 12, 15, 8, 12, 20]
        )
        
        self.ws = original_ws
    
    def _create_custom_quality_section(self, stats, parametros):
        """Create custom quality section."""
        self._create_section_title('A8', "Análisis Personalizado de Calidad")
        
        custom_metrics = []
        if parametros.get('include_dimensions', True):
            custom_metrics.extend([
                [EXCEL_AVG_ALTO, f"{stats['avg_dimensions']['alto']} mm"],
                [EXCEL_AVG_ANCHO, f"{stats['avg_dimensions']['ancho']} mm"],
                [EXCEL_AVG_GROSOR, f"{stats['avg_dimensions']['grosor']} mm"],
            ])
        if parametros.get('include_weight', True):
            custom_metrics.append([EXCEL_AVG_PESO, f"{stats['avg_weight']} g"])
        if parametros.get('include_confidence', True):
            custom_metrics.append([EXCEL_AVG_CONFIDENCE, f"{stats['avg_confidence']}%"])
        
        data = [[EXCEL_COL_METRIC, EXCEL_COL_VALUE]] + custom_metrics
        self._create_table_with_data(
            data,
            start_row=10,
            header_row=10,
            column_widths={'A': 20, 'B': 15}
        )
    
    def _create_custom_finca_section(self, finca, parametros):
        """Create custom farm section."""
        self._create_section_title('A8', f"Análisis Personalizado - {finca.nombre}")
        
        custom_metrics = []
        if parametros.get('include_basic_info', True):
            custom_metrics.extend([
                ['Nombre', finca.nombre],
                ['Ubicación', finca.ubicacion_completa],
                ['Hectáreas', f"{finca.hectareas} ha"],
            ])
        if parametros.get('include_lotes', True):
            custom_metrics.extend([
                [EXCEL_TOTAL_LOTES, str(finca.total_lotes)],
                [EXCEL_LOTES_ACTIVOS, str(finca.lotes_activos)],
            ])
        if parametros.get('include_quality', True):
            custom_metrics.extend([
                [EXCEL_TOTAL_ANALISIS, str(finca.total_analisis)],
                ['Calidad Promedio', f"{finca.calidad_promedio}%"],
            ])
        
        data = [['Campo', 'Valor']] + custom_metrics
        self._create_table_with_data(
            data,
            start_row=10,
            header_row=10,
            column_widths={'A': 20, 'B': 30},
            body_alignment='left'
        )
    
    def _create_custom_audit_section(self, stats, parametros):
        """Create custom audit section."""
        self._create_section_title('A8', "Análisis Personalizado de Auditoría")
        
        custom_metrics = []
        if parametros.get('include_activity', True):
            custom_metrics.extend([
                ['Total de Actividades', str(stats['total_activities'])],
                ['Actividades Hoy', str(stats['activities_today'])],
            ])
        if parametros.get('include_top_users', True):
            custom_metrics.append(['Usuarios Más Activos', str(len(stats['top_users']))])
        if parametros.get('include_action_types', True):
            custom_metrics.append(['Tipos de Acción', str(len(stats['activities_by_action']))])
        
        data = [[EXCEL_COL_METRIC, EXCEL_COL_VALUE]] + custom_metrics
        self._create_table_with_data(
            data,
            start_row=10,
            header_row=10,
            column_widths={'A': 20, 'B': 15}
        )

