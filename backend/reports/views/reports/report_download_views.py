"""
Download views for reports in CacaoScan.
Handles downloading generated reports and special admin reports.
"""
import logging
import json
import traceback
from datetime import datetime
from django.utils import timezone
from django.http import HttpResponse, FileResponse
from django.contrib.auth import get_user_model
from rest_framework.permissions import IsAuthenticated, IsAdminUser
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from drf_yasg.utils import swagger_auto_schema
from drf_yasg import openapi
from django.utils.encoding import escape_uri_path

from reports.models import ReporteGenerado
from api.utils.model_imports import get_models_safely
from reports.services import ExcelAgricultoresGenerator, ExcelUsuariosGenerator
from api.serializers import ErrorResponseSerializer
from .report_crud_views import ExcelRenderer

# Import models safely
models = get_models_safely({
    'Finca': 'fincas_app.models.Finca'
})
Finca = models['Finca']

User = get_user_model()
logger = logging.getLogger("cacaoscan.api")

# Content type constants
CONTENT_TYPE_JSON = 'application/json'


class ReporteDownloadView(APIView):
    """
    View for downloading a generated report.
    """
    permission_classes = [IsAuthenticated]
    
    @swagger_auto_schema(
        operation_description="Descarga un reporte generado",
        operation_summary="Descargar reporte",
        responses={
            200: openapi.Response(description="Archivo descargado exitosamente"),
            404: ErrorResponseSerializer,
            410: ErrorResponseSerializer,
            401: ErrorResponseSerializer,
        },
        tags=['Reportes']
    )
    def get(self, request, reporte_id):
        """Download report."""
        try:
            reporte = ReporteGenerado.objects.get(id=reporte_id, usuario=request.user)
            
            # Verify state (estado is a ForeignKey to Parametro)
            estado_codigo = reporte.estado.codigo if hasattr(reporte.estado, 'codigo') else None
            if estado_codigo != 'COMPLETADO':
                return Response({
                    'error': 'El reporte aún no está listo para descarga',
                    'details': f'El reporte está en estado: {estado_codigo or "desconocido"}'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Verify if expired
            if reporte.esta_expirado:
                return Response({
                    'error': 'El reporte ha expirado y ya no está disponible',
                    'details': 'El reporte ha expirado y ya no está disponible'
                }, status=status.HTTP_410_GONE)
            
            # Verify file exists
            if not reporte.archivo:
                return Response({
                    'error': 'El archivo del reporte no está disponible',
                    'details': 'El archivo del reporte no está disponible'
                }, status=status.HTTP_404_NOT_FOUND)
            
            # Get filename and format extension
            formato_codigo = reporte.formato.codigo if hasattr(reporte.formato, 'codigo') else 'PDF'
            if reporte.nombre_archivo:
                filename = reporte.nombre_archivo
            else:
                # Generate filename with correct extension
                extension_map = {
                    'PDF': 'pdf',
                    'EXCEL': 'xlsx',
                    'CSV': 'csv',
                    'JSON': 'json'
                }
                extension = extension_map.get(formato_codigo, 'pdf')
                filename = f"{reporte.titulo}.{extension}"
            
            # Prepare download response
            response = FileResponse(
                reporte.archivo,
                as_attachment=True,
                filename=filename
            )
            
            # Configure headers according to format
            if formato_codigo == 'PDF':
                response['Content-Type'] = 'application/pdf'
            elif formato_codigo == 'EXCEL':
                response['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            elif formato_codigo == 'CSV':
                response['Content-Type'] = 'text/csv'
            elif formato_codigo == 'JSON':
                response['Content-Type'] = CONTENT_TYPE_JSON
            
            # Ensure Content-Disposition header is set correctly
            response['Content-Disposition'] = f'attachment; filename="{escape_uri_path(filename)}"'
            
            logger.info(f"Reporte {reporte_id} descargado por usuario {request.user.username}")
            
            return response
            
        except ReporteGenerado.DoesNotExist:
            return Response({
                'error': 'Reporte no encontrado',
                'details': 'El reporte solicitado no existe o no tienes permiso para acceder a él'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"Error descargando reporte {reporte_id}: {e}")
            return Response({
                'error': 'Error interno del servidor',
                'details': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class ReporteAgricultoresView(APIView):
    """
    Generates an Excel file with farmer and farm information.
    Requires JWT authentication and admin permissions.
    """
    permission_classes = [IsAuthenticated, IsAdminUser]
    renderer_classes = [ExcelRenderer]
    
    def _get_excel_constants(self):
        """Get Excel styling constants."""
        from openpyxl.styles import Border, Side
        return {
            'COLOR_VERDE_PRIMITIVO': "166534",
            'COLOR_VERDE_CLARO': "A7F3D0",
            'COLOR_GRIS_SUAVE': "6B7280",
            'thin_border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
    
    def _setup_workbook_title(self, ws, constants):
        """Setup workbook title and headers."""
        from openpyxl.styles import Font, PatternFill, Alignment
        
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
        title_cell = ws.cell(row=1, column=1, value="Reporte de Agricultores y Fincas - CacaoScan")
        title_cell.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color=constants['COLOR_VERDE_PRIMITIVO'], end_color=constants['COLOR_VERDE_PRIMITIVO'], fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 10
        
        headers = [
            'Agricultor', 'Email', 'Telefono', 'Departamento', 'Municipio',
            'Finca', 'Hectareas', 'Estado Finca', 'Fecha Registro Finca'
        ]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = Font(name="Calibri", size=12, bold=True, color=constants['COLOR_VERDE_PRIMITIVO'])
            cell.fill = PatternFill(start_color=constants['COLOR_VERDE_CLARO'], end_color=constants['COLOR_VERDE_CLARO'], fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = constants['thin_border']
        ws.row_dimensions[3].height = 25
    
    def _get_farmers_with_farms(self):
        """Get list of farmers that have farms."""
        agricultores_ids = Finca.objects.values_list("agricultor_id", flat=True).distinct()
        agricultores_ids_list = list(agricultores_ids)
        
        logger.debug("[DEBUG] Encontrados %d agricultores con fincas", len(agricultores_ids_list))
        
        if not agricultores_ids_list:
            logger.warning("[WARNING] No hay agricultores con fincas en la base de datos")
            return []
        
        agricultores = User.objects.filter(id__in=agricultores_ids_list).prefetch_related('fincas_app_fincas')
        agricultores_list = list(agricultores)
        logger.info("[INFO] Obtenidos %d agricultores con prefetch", len(agricultores_list))
        return agricultores_list
    
    def _add_no_data_row(self, ws, constants):
        """Add row indicating no data available."""
        from openpyxl.styles import Font, Alignment
        no_data_row = ["-", "Sin datos", "-", "-", "-", "-", "-", "-", "-"]
        for col_idx, value in enumerate(no_data_row, start=1):
            cell = ws.cell(row=4, column=col_idx, value=value)
            cell.font = Font(name="Calibri", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = constants['thin_border']
    
    def _get_farmer_info(self, agricultor):
        """Extract basic farmer information."""
        nombre = agricultor.get_full_name() or agricultor.username or f"Usuario {agricultor.id}"
        email = agricultor.email or ""
        
        telefono = ""
        try:
            persona = getattr(agricultor, 'persona', None)
            if persona and hasattr(persona, 'telefono') and persona.telefono:
                telefono = str(persona.telefono)
        except Exception:
            pass
        
        return nombre, email, telefono
    
    def _get_farmer_farms(self, agricultor):
        """Get farmer's farms safely."""
        fincas_list = []
        try:
            if hasattr(agricultor, 'fincas_app_fincas'):
                fincas_queryset = agricultor.fincas_app_fincas.all()
                fincas_list = list(fincas_queryset)
        except Exception as finca_error:
            logger.warning("[WARNING] Error obteniendo fincas para usuario %d: %s", agricultor.id, finca_error)
        return fincas_list
    
    def _extract_finca_data(self, finca):
        """Extract farm data safely."""
        depto = str(finca.departamento) if finca.departamento else ""
        municipio = str(finca.municipio) if finca.municipio else ""
        finca_nombre = str(finca.nombre) if finca.nombre else "Sin nombre"
        
        hectareas_val = 0.0
        try:
            if finca.hectareas is not None:
                hectareas_val = float(finca.hectareas)
        except (ValueError, TypeError, AttributeError):
            hectareas_val = 0.0
        
        estado_finca = "Activa" if (hasattr(finca, 'activa') and finca.activa) else "Inactiva"
        
        fecha_registro_str = ""
        try:
            if hasattr(finca, 'fecha_registro') and finca.fecha_registro:
                fecha_registro_str = finca.fecha_registro.strftime('%d/%m/%Y')
        except Exception:
            pass
        
        return {
            'depto': depto,
            'municipio': municipio,
            'nombre': finca_nombre,
            'hectareas': hectareas_val,
            'estado': estado_finca,
            'fecha_registro': fecha_registro_str
        }
    
    def _add_data_row(self, ws, row_data, row_num, constants):
        """Add a data row to Excel with styling."""
        from openpyxl.styles import Font, Alignment
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.font = Font(name="Calibri", size=11)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = constants['thin_border']
            
            if col_idx == 7 and isinstance(value, (int, float)):
                cell.alignment = Alignment(horizontal="right", vertical="center")
    
    def _add_farm_row(self, ws, nombre, email, telefono, finca_data, row_num, constants):
        """Add a row for farmer with farm data."""
        row_data = [
            nombre, email, telefono, finca_data['depto'], finca_data['municipio'],
            finca_data['nombre'], finca_data['hectareas'], finca_data['estado'], finca_data['fecha_registro']
        ]
        self._add_data_row(ws, row_data, row_num, constants)
    
    def _add_farmer_only_row(self, ws, nombre, email, telefono, agricultor, row_num, constants):
        """Add row for farmer without farms."""
        fecha_registro_str = ""
        try:
            if agricultor.date_joined:
                fecha_registro_str = agricultor.date_joined.strftime('%d/%m/%Y')
        except Exception:
            pass
        
        row_data = [
            nombre, email, telefono,
            "", "", "", "", "",
            fecha_registro_str
        ]
        self._add_data_row(ws, row_data, row_num, constants)
    
    def _process_farmer_rows(self, ws, agricultor, constants, current_row):
        """Process a single farmer and add rows to Excel."""
        nombre, email, telefono = self._get_farmer_info(agricultor)
        fincas_list = self._get_farmer_farms(agricultor)
        
        rows_added = 0
        
        if fincas_list:
            for finca in fincas_list:
                try:
                    finca_data = self._extract_finca_data(finca)
                    self._add_farm_row(ws, nombre, email, telefono, finca_data, current_row, constants)
                    current_row += 1
                    rows_added += 1
                except Exception as finca_row_error:
                    logger.warning("[WARNING] Error procesando finca %s: %s", 
                                 getattr(finca, 'id', 'unknown'), finca_row_error)
        else:
            self._add_farmer_only_row(ws, nombre, email, telefono, agricultor, current_row, constants)
            current_row += 1
            rows_added += 1
        
        return current_row, rows_added
    
    def _add_footer(self, ws, current_row, request_user, constants):
        """Add footer with user and date information."""
        from openpyxl.styles import Font, Alignment
        
        footer_start_row = current_row + 2
        ws.row_dimensions[footer_start_row - 1].height = 10
        
        ws.merge_cells(start_row=footer_start_row, start_column=1, end_row=footer_start_row, end_column=9)
        generated_cell = ws.cell(row=footer_start_row, column=1, value=f"Generado por: {request_user.username}")
        generated_cell.font = Font(name="Calibri", size=10, color=constants['COLOR_GRIS_SUAVE'])
        generated_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        now = datetime.now()
        fecha_generacion = now.strftime('%d/%m/%Y %I:%M:%S %p')
        ws.merge_cells(start_row=footer_start_row + 1, start_column=1, end_row=footer_start_row + 1, end_column=9)
        fecha_cell = ws.cell(row=footer_start_row + 1, column=1, value=f"Fecha: {fecha_generacion}")
        fecha_cell.font = Font(name="Calibri", size=10, color=constants['COLOR_GRIS_SUAVE'])
        fecha_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    def _adjust_column_widths(self, ws):
        """Adjust column widths."""
        column_widths = {
            'A': 25, 'B': 30, 'C': 15, 'D': 18, 'E': 18,
            'F': 20, 'G': 12, 'H': 15, 'I': 18,
        }
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
    
    def _save_and_validate_excel(self, wb):
        """Save workbook to memory and validate content."""
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        excel_content = output.getvalue()
        
        if not excel_content or len(excel_content) < 50:
            output.close()
            logger.error("[ERROR] El archivo Excel generado está vacío o demasiado pequeño: %d bytes", 
                       len(excel_content) if excel_content else 0)
            return None, None
        
        return excel_content, output
    
    def _create_excel_response(self, excel_content, output, request_user):
        """Create HTTP response with Excel file."""
        fecha_actual = timezone.now().strftime("%Y-%m-%d_%H-%M-%S")
        filename = f"reporte_agricultores_{fecha_actual}.xlsx"
        
        response = HttpResponse(
            excel_content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{escape_uri_path(filename)}"'
        response["Content-Length"] = str(output.getbuffer().nbytes)
        output.close()
        
        return response
    
    def get(self, request, *args, **kwargs):
        """
        Generates and downloads an Excel file with information of all farmers and their farms.
        """
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        
        try:
            logger.info("[INFO] Generando reporte de agricultores para %s", request.user.username)
            
            # Validate that Finca model is available
            if Finca is None:
                raise ImportError("El modelo Finca no está disponible")
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Agricultores"
            
            constants = self._get_excel_constants()
            self._setup_workbook_title(ws, constants)
            
            try:
                agricultores_list = self._get_farmers_with_farms()
                if not agricultores_list:
                    self._add_no_data_row(ws, constants)
            except Exception as query_error:
                logger.error("[ERROR] Error obteniendo agricultores: %s", query_error, exc_info=True)
                raise
            
            current_row = 4
            rows_added = 0
            
            for agricultor in agricultores_list:
                try:
                    current_row, row_count = self._process_farmer_rows(ws, agricultor, constants, current_row)
                    rows_added += row_count
                except Exception as agricultor_error:
                    logger.warning("[WARNING] Error procesando agricultor %s: %s", 
                                 getattr(agricultor, 'id', 'unknown'), agricultor_error)
                    continue
            
            logger.info("[INFO] Procesados %d filas de datos", rows_added)
            
            self._add_footer(ws, current_row, request.user, constants)
            self._adjust_column_widths(ws)
            
            excel_content, output = self._save_and_validate_excel(wb)
            if excel_content is None:
                return HttpResponse(
                    "Error interno al generar el reporte Excel: contenido inválido",
                    status=500,
                    content_type='text/plain'
                )
            
            logger.info("[SUCCESS] Reporte generado correctamente (%d bytes) para usuario %s", 
                       len(excel_content), request.user.username)
            
            return self._create_excel_response(excel_content, output, request.user)
            
        except Exception as e:
            error_trace = traceback.format_exc()
            logger.error("[ERROR] Error generando reporte de agricultores: %s\n%s", str(e), error_trace)
            
            return HttpResponse(
                f"Error interno al generar el reporte: {str(e)}",
                status=500,
                content_type='text/plain'
            )


class ReporteUsuariosView(APIView):
    """
    View for generating and downloading Excel report of system users.
    Only accessible to administrators.
    """
    permission_classes = [IsAuthenticated, IsAdminUser]
    # Disable DRF renderers to return binary response directly
    renderer_classes = []
    
    @swagger_auto_schema(
        operation_description="Genera y descarga un archivo Excel con información de todos los usuarios del sistema",
        operation_summary="Reporte de usuarios (Excel)",
        responses={
            200: openapi.Response(description="Archivo Excel generado exitosamente"),
            401: ErrorResponseSerializer,
            403: ErrorResponseSerializer,
        },
        tags=['Reportes']
    )
    def get(self, request):
        try:
            # Generate Excel report
            excel_service = ExcelUsuariosGenerator()
            excel_content = excel_service.generate_users_report()
            
            # Validate content is not empty
            if not excel_content or len(excel_content) < 100:
                logger.error("El contenido del reporte Excel está vacío o corrupto")
                return HttpResponse(
                    json.dumps({'error': 'Error al generar el reporte Excel: contenido vacío'}),
                    content_type=CONTENT_TYPE_JSON,
                    status=500
                )
            
            # Create HTTP response with file (use HttpResponse directly, not DRF Response)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'reporte_usuarios_{timestamp}.xlsx'
            
            response = HttpResponse(
                excel_content,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
            # Configure file name with correct encoding
            response['Content-Disposition'] = f'attachment; filename="{escape_uri_path(filename)}"'
            response['Content-Length'] = len(excel_content)
            
            logger.info(f"Reporte de usuarios generado por {request.user.username} - Tamaño: {len(excel_content)} bytes")
            
            return response
            
        except Exception as e:
            logger.error(f"Error generando reporte de usuarios: {e}", exc_info=True)
            # For errors, use HttpResponse with JSON instead of DRF Response
            error_response = HttpResponse(
                json.dumps({
                    'error': 'Error al generar el reporte de usuarios',
                    'detail': str(e)
                }),
                content_type=CONTENT_TYPE_JSON,
                status=500
            )
            return error_response

