"""
Generador de reportes Excel para CacaoScan.
"""
import logging
import io
from datetime import datetime
from django.utils import timezone
from django.db.models import Count, Avg, Sum
from django.http import HttpResponse
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference, PieChart

# Importar desde apps modulares
try:
    from notifications.models import Notification
except ImportError:
    Notification = None

try:
    from images_app.models import CacaoImage, CacaoPrediction
except ImportError:
    CacaoImage = None
    CacaoPrediction = None

try:
    from fincas_app.models import Finca, Lote
except ImportError:
    Finca = None
    Lote = None

try:
    from audit.models import ActivityLog
except ImportError:
    ActivityLog = None

from .models import LoginHistory, ReporteGenerado

logger = logging.getLogger("cacaoscan.api")


class CacaoReportExcelGenerator:
    """
    Generador de reportes Excel avanzado para CacaoScan.
    """
    
    def __init__(self):
        self.workbook = None
        self.ws = None
    
    def generate_farmers_report(self):
        """
        Generar reporte Excel de agricultores con sus fincas.
        
        Returns:
            bytes: Contenido del archivo Excel
        """
        try:
            from django.contrib.auth.models import User
            
            self.workbook = Workbook()
            self.ws = self.workbook.active
            self.ws.title = "Agricultores"
            
            # Configurar columnas
            columns = [
                'Agricultor', 'Email', 'Teléfono', 'Departamento', 'Municipio',
                'Finca', 'Hectáreas', 'Estado Finca', 'Fecha Registro Finca'
            ]
            self.ws.append(columns)
            
            # Estilo de encabezado
            header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in self.ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Obtener todos los usuarios que no son superusuarios (agricultores y analistas)
            farmers = User.objects.filter(is_superuser=False).select_related('auth_profile').prefetch_related('api_fincas')
            
            # Contador de filas
            row_num = 2
            
            for farmer in farmers:
                # Información del agricultor
                name = f"{farmer.first_name} {farmer.last_name}".strip() or farmer.username
                email = farmer.email
                phone = getattr(farmer, 'auth_profile', None) and getattr(farmer.auth_profile, 'phone_number', '') or ''
                
                # Obtener fincas del agricultor
                fincas = farmer.api_fincas.all()
                
                if fincas.exists():
                    # Si tiene fincas, crear una fila por cada finca
                    for finca in fincas:
                        self.ws.append([
                            name,
                            email,
                            phone,
                            finca.departamento,
                            finca.municipio,
                            finca.nombre,
                            float(finca.hectareas),
                            "Activa" if finca.activa else "Inactiva",
                            finca.fecha_registro.strftime('%Y-%m-%d') if finca.fecha_registro else ''
                        ])
                        row_num += 1
                else:
                    # Si no tiene fincas, agregar fila con campos vacíos para finca
                    self.ws.append([
                        name,
                        email,
                        phone,
                        '',
                        '',
                        '',
                        '',
                        '',
                        farmer.date_joined.strftime('%Y-%m-%d') if farmer.date_joined else ''
                    ])
                    row_num += 1
            
            # Ajustar ancho de columnas
            column_widths = {
                'A': 25,  # Agricultor
                'B': 30,  # Email
                'C': 15,  # Teléfono
                'D': 20,  # Departamento
                'E': 20,  # Municipio
                'F': 20,  # Finca
                'G': 12,  # Hectáreas
                'H': 15,  # Estado
                'I': 18,  # Fecha Registro
            }
            
            for col, width in column_widths.items():
                self.ws.column_dimensions[col].width = width
            
            # Centrar encabezados
            for row in self.ws.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Guardar en buffer
            buffer = io.BytesIO()
            self.workbook.save(buffer)
            buffer.seek(0)
            return buffer.getvalue()
            
        except Exception as e:
            logger.error(f"Error generando reporte Excel de agricultores: {e}", exc_info=True)
            raise
    
    def generate_users_report(self):
        """
        Generar reporte Excel profesional de usuarios con sus fincas asociadas.
        Incluye formato visual profesional (colores, bordes, alineación).
        Si no hay usuarios, muestra mensaje "Sin registros disponibles".
        
        Returns:
            bytes: Contenido del archivo Excel
        """
        try:
            from django.contrib.auth.models import User
            from .models import Finca
            from openpyxl.styles import Border, Side
            
            self.workbook = Workbook()
            self.ws = self.workbook.active
            self.ws.title = "Usuarios y Fincas"
            
            # === Estilos básicos ===
            bold_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            align_center = Alignment(horizontal='center', vertical='center')
            align_vertical = Alignment(vertical='center')
            
            # === Encabezados ===
            headers = [
                'ID Usuario', 'Nombre', 'Correo', 'Rol', 'Activo', 'Fecha Registro',
                'Finca', 'Departamento', 'Municipio', 'Área (ha)', 'Latitud', 'Longitud'
            ]
            self.ws.append(headers)
            
            # Aplicar estilo a encabezados
            for col in self.ws[1]:
                col.font = bold_font
                col.fill = header_fill
                col.alignment = align_center
                col.border = thin_border
            
            # Obtener todos los usuarios con prefetch de fincas
            users = User.objects.all().order_by('-date_joined').select_related('auth_profile', 'auth_email_token').prefetch_related('api_fincas', 'groups')
            
            if users.exists():
                # Iterar usuarios y agregar información de fincas
                for user in users:
                    fincas = user.api_fincas.all()
                    
                    # Determinar rol del usuario
                    if user.is_superuser or user.is_staff:
                        rol = 'admin'
                    elif user.groups.filter(name='analyst').exists():
                        rol = 'analyst'
                    else:
                        rol = 'farmer'
                    
                    if fincas.exists():
                        # Si el usuario tiene fincas, crear una fila por cada finca
                        for finca in fincas:
                            self.ws.append([
                                user.id,
                                f"{user.first_name} {user.last_name}".strip() or user.username,
                                user.email,
                                rol,
                                'Sí' if user.is_active else 'No',
                                user.date_joined.strftime('%Y-%m-%d'),
                                finca.nombre,
                                finca.departamento or '—',
                                finca.municipio or '—',
                                float(finca.hectareas),
                                float(finca.coordenadas_lat) if finca.coordenadas_lat is not None else '—',
                                float(finca.coordenadas_lng) if finca.coordenadas_lng is not None else '—',
                            ])
                    else:
                        # Si no tiene fincas, agregar fila con "Sin fincas"
                        self.ws.append([
                            user.id,
                            f"{user.first_name} {user.last_name}".strip() or user.username,
                            user.email,
                            rol,
                            'Sí' if user.is_active else 'No',
                            user.date_joined.strftime('%Y-%m-%d'),
                            'Sin fincas',
                            '—', '—', '—', '—', '—'
                        ])
                
                # Aplicar bordes a todas las celdas con datos
                for row in self.ws.iter_rows(min_row=2, max_row=self.ws.max_row, max_col=len(headers)):
                    for cell in row:
                        cell.border = thin_border
                        cell.alignment = align_vertical
                
                # Ajustar ancho de columnas automáticamente
                for col in self.ws.columns:
                    max_length = 0
                    for cell in col:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)  # Máximo 50 caracteres
                    self.ws.column_dimensions[col[0].column_letter].width = adjusted_width
                
            else:
                # Si no hay usuarios, mostrar mensaje amigable
                self.ws.merge_cells('A1:L2')
                cell = self.ws['A1']
                cell.value = "Sin registros disponibles"
                cell.font = Font(bold=True, size=14, color="FF0000")
                cell.alignment = align_center
                cell.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
                self.ws.row_dimensions[1].height = 40
            
            # Guardar en buffer de memoria (binario)
            buffer = io.BytesIO()
            self.workbook.save(buffer)
            buffer.seek(0)
            
            # Validar que el buffer no esté vacío
            content = buffer.getvalue()
            if not content or len(content) < 100:
                raise ValueError("El archivo Excel generado está vacío o corrupto")
            
            logger.info(f"Reporte Excel de usuarios y fincas generado correctamente ({len(content)} bytes)")
            return content
            
        except Exception as e:
            logger.error(f"Error generando reporte Excel de usuarios: {e}", exc_info=True)
            raise
    
    def generate_quality_report(self, user, filtros=None):
        """
        Generar reporte de calidad en Excel.
        
        Args:
            user: Usuario que solicita el reporte
            filtros: Filtros a aplicar
        """
        try:
            self.workbook = Workbook()
            self.ws = self.workbook.active
            self.ws.title = "Reporte de Calidad"
            
            # Aplicar filtros
            queryset = self._apply_filters(CacaoPrediction.objects.all(), filtros)
            
            # Crear encabezado
            self._create_header("Reporte de Calidad de Granos de Cacao", user)
            
            # Estadísticas generales
            stats = self._get_quality_stats(queryset)
            self._create_stats_section(stats)
            
            # Tabla de análisis detallados
            self._create_detailed_analyses_table(queryset)
            
            # Gráfico de distribución de calidad
            self._create_quality_chart(stats)
            
            # Hoja de resumen
            self._create_summary_sheet(stats, user)
            
            # Guardar en buffer
            buffer = io.BytesIO()
            self.workbook.save(buffer)
            buffer.seek(0)
            return buffer.getvalue()
            
        except Exception as e:
            logger.error(f"Error generando reporte Excel de calidad: {e}")
            raise
    
    def generate_finca_report(self, finca_id, user, filtros=None):
        """
        Generar reporte de finca en Excel.
        
        Args:
            finca_id: ID de la finca
            user: Usuario que solicita el reporte
            filtros: Filtros a aplicar
        """
        try:
            finca = Finca.objects.get(id=finca_id)
            self.workbook = Workbook()
            self.ws = self.workbook.active
            self.ws.title = f"Finca {finca.nombre}"
            
            # Crear encabezado
            self._create_header(f"Reporte de Finca: {finca.nombre}", user)
            
            # Información de la finca
            self._create_finca_info_section(finca)
            
            # Estadísticas de lotes
            lotes_stats = self._get_lotes_stats(finca)
            self._create_lotes_stats_section(lotes_stats)
            
            # Análisis por lote
            self._create_lotes_analysis_section(finca)
            
            # Hoja de análisis detallados
            self._create_detailed_lotes_sheet(finca)
            
            # Guardar en buffer
            buffer = io.BytesIO()
            self.workbook.save(buffer)
            buffer.seek(0)
            return buffer.getvalue()
            
        except Exception as e:
            logger.error(f"Error generando reporte Excel de finca: {e}")
            raise
    
    def generate_audit_report(self, user, filtros=None):
        """
        Generar reporte de auditoría en Excel.
        
        Args:
            user: Usuario que solicita el reporte
            filtros: Filtros a aplicar
        """
        try:
            self.workbook = Workbook()
            self.ws = self.workbook.active
            self.ws.title = "Auditoría"
            
            # Crear encabezado
            self._create_header("Reporte de Auditoría del Sistema", user)
            
            # Estadísticas de actividad
            activity_stats = self._get_activity_stats(filtros)
            self._create_activity_stats_section(activity_stats)
            
            # Estadísticas de logins
            login_stats = self._get_login_stats(filtros)
            self._create_login_stats_section(login_stats)
            
            # Hoja de actividades detalladas
            self._create_detailed_activities_sheet(filtros)
            
            # Hoja de logins detallados
            self._create_detailed_logins_sheet(filtros)
            
            # Guardar en buffer
            buffer = io.BytesIO()
            self.workbook.save(buffer)
            buffer.seek(0)
            return buffer.getvalue()
            
        except Exception as e:
            logger.error(f"Error generando reporte Excel de auditoría: {e}")
            raise
    
    def generate_custom_report(self, user, tipo_reporte, parametros, filtros=None):
        """
        Generar reporte personalizado en Excel.
        
        Args:
            user: Usuario que solicita el reporte
            tipo_reporte: Tipo de reporte
            parametros: Parámetros del reporte
            filtros: Filtros a aplicar
        """
        try:
            self.workbook = Workbook()
            self.ws = self.workbook.active
            self.ws.title = "Reporte Personalizado"
            
            # Crear encabezado
            self._create_header(f"Reporte Personalizado: {tipo_reporte}", user)
            
            # Generar según tipo
            if tipo_reporte == 'calidad':
                queryset = self._apply_filters(CacaoPrediction.objects.all(), filtros)
                stats = self._get_quality_stats(queryset)
                self._create_custom_quality_section(stats, parametros)
            elif tipo_reporte == 'finca':
                if parametros.get('finca_id'):
                    finca = Finca.objects.get(id=parametros['finca_id'])
                    self._create_custom_finca_section(finca, parametros)
            elif tipo_reporte == 'auditoria':
                activity_stats = self._get_activity_stats(filtros)
                self._create_custom_audit_section(activity_stats, parametros)
            
            # Guardar en buffer
            buffer = io.BytesIO()
            self.workbook.save(buffer)
            buffer.seek(0)
            return buffer.getvalue()
            
        except Exception as e:
            logger.error(f"Error generando reporte Excel personalizado: {e}")
            raise
    
    def _apply_filters(self, queryset, filtros):
        """Aplicar filtros al queryset."""
        if not filtros:
            return queryset
        
        # Filtro por fecha
        if filtros.get('fecha_desde'):
            queryset = queryset.filter(created_at__date__gte=filtros['fecha_desde'])
        if filtros.get('fecha_hasta'):
            queryset = queryset.filter(created_at__date__lte=filtros['fecha_hasta'])
        
        # Filtro por usuario
        if filtros.get('usuario_id'):
            queryset = queryset.filter(image__user_id=filtros['usuario_id'])
        
        # Filtro por finca
        if filtros.get('finca_id'):
            queryset = queryset.filter(image__finca=filtros['finca_id'])
        
        # Filtro por lote
        if filtros.get('lote_id'):
            queryset = queryset.filter(image__lote_id=filtros['lote_id'])
        
        return queryset
    
    def _create_header(self, title, user):
        """Crear encabezado del reporte."""
        # Título principal
        self.ws['A1'] = title
        self.ws['A1'].font = Font(size=16, bold=True, color="2F4F4F")
        self.ws['A1'].alignment = Alignment(horizontal='center')
        self.ws.merge_cells('A1:F1')
        
        # Información del reporte
        self.ws['A3'] = f"Generado el: {timezone.now().strftime('%d/%m/%Y %H:%M')}"
        self.ws['A3'].font = Font(size=10, italic=True)
        
        self.ws['A4'] = f"Usuario: {user.get_full_name() or user.username}"
        self.ws['A4'].font = Font(size=10, italic=True)
        
        # Espacio
        self.ws['A6'] = ""
    
    def _get_quality_stats(self, queryset):
        """Obtener estadísticas de calidad."""
        total_analyses = queryset.count()
        
        if total_analyses == 0:
            return {
                'total_analyses': 0,
                'avg_confidence': 0,
                'quality_distribution': {},
                'avg_dimensions': {},
                'avg_weight': 0
            }
        
        # Estadísticas de confianza
        avg_confidence = queryset.aggregate(avg=Avg('average_confidence'))['avg'] or 0
        
        # Distribución de calidad
        quality_distribution = {
            'Excelente (≥90%)': queryset.filter(average_confidence__gte=0.9).count(),
            'Buena (80-89%)': queryset.filter(average_confidence__gte=0.8, average_confidence__lt=0.9).count(),
            'Regular (70-79%)': queryset.filter(average_confidence__gte=0.7, average_confidence__lt=0.8).count(),
            'Baja (<70%)': queryset.filter(average_confidence__lt=0.7).count(),
        }
        
        # Dimensiones promedio
        avg_dimensions = queryset.aggregate(
            avg_alto=Avg('alto_mm'),
            avg_ancho=Avg('ancho_mm'),
            avg_grosor=Avg('grosor_mm')
        )
        
        # Peso promedio
        avg_weight = queryset.aggregate(avg=Avg('peso_g'))['avg'] or 0
        
        return {
            'total_analyses': total_analyses,
            'avg_confidence': round(float(avg_confidence) * 100, 2),
            'quality_distribution': quality_distribution,
            'avg_dimensions': {
                'alto': round(float(avg_dimensions['avg_alto'] or 0), 2),
                'ancho': round(float(avg_dimensions['avg_ancho'] or 0), 2),
                'grosor': round(float(avg_dimensions['avg_grosor'] or 0), 2),
            },
            'avg_weight': round(float(avg_weight), 2)
        }
    
    def _create_stats_section(self, stats):
        """Crear sección de estadísticas."""
        # Título de sección
        self.ws['A8'] = "Estadísticas Generales"
        self.ws['A8'].font = Font(size=14, bold=True, color="2F4F4F")
        
        # Datos de estadísticas
        data = [
            ['Métrica', 'Valor'],
            ['Total de Análisis', stats['total_analyses']],
            ['Confianza Promedio', f"{stats['avg_confidence']}%"],
            ['Alto Promedio', f"{stats['avg_dimensions']['alto']} mm"],
            ['Ancho Promedio', f"{stats['avg_dimensions']['ancho']} mm"],
            ['Grosor Promedio', f"{stats['avg_dimensions']['grosor']} mm"],
            ['Peso Promedio', f"{stats['avg_weight']} g"],
        ]
        
        # Crear tabla
        for row_num, row_data in enumerate(data, 10):
            for col_num, cell_value in enumerate(row_data, 1):
                cell = self.ws.cell(row=row_num, column=col_num, value=cell_value)
                
                # Estilo para encabezados
                if row_num == 10:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center')
                else:
                    cell.alignment = Alignment(horizontal='center')
                
                # Bordes
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # Ajustar ancho de columnas
        self.ws.column_dimensions['A'].width = 20
        self.ws.column_dimensions['B'].width = 15
    
    def _create_detailed_analyses_table(self, queryset):
        """Crear tabla de análisis detallados."""
        # Título de sección
        self.ws['A18'] = "Análisis Detallados"
        self.ws['A18'].font = Font(size=14, bold=True, color="2F4F4F")
        
        # Encabezados
        headers = ['ID', 'Usuario', 'Finca', 'Región', 'Fecha', 'Alto (mm)', 'Ancho (mm)', 'Grosor (mm)', 'Peso (g)', 'Confianza']
        for col_num, header in enumerate(headers, 1):
            cell = self.ws.cell(row=20, column=col_num, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Datos (limitar a 100 registros para evitar archivos muy grandes)
        analyses = queryset.select_related('image', 'image__user').order_by('-created_at')[:100]
        
        for row_num, analysis in enumerate(analyses, 21):
            data = [
                analysis.id,
                analysis.image.user.username,
                analysis.image.finca or 'N/A',
                analysis.image.region or 'N/A',
                analysis.created_at.strftime('%d/%m/%Y %H:%M'),
                round(float(analysis.alto_mm), 2),
                round(float(analysis.ancho_mm), 2),
                round(float(analysis.grosor_mm), 2),
                round(float(analysis.peso_g), 2),
                f"{analysis.average_confidence:.2%}"
            ]
            
            for col_num, value in enumerate(data, 1):
                cell = self.ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(horizontal='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # Ajustar ancho de columnas
        column_widths = [8, 12, 15, 12, 18, 10, 10, 10, 10, 12]
        for i, width in enumerate(column_widths, 1):
            self.ws.column_dimensions[chr(64 + i)].width = width
    
    def _create_quality_chart(self, stats):
        """Crear gráfico de distribución de calidad."""
        if not stats['quality_distribution']:
            return
        
        # Crear nueva hoja para el gráfico
        chart_ws = self.workbook.create_sheet("Distribución de Calidad")
        
        # Datos para el gráfico
        chart_ws['A1'] = "Categoría"
        chart_ws['B1'] = "Cantidad"
        chart_ws['C1'] = "Porcentaje"
        
        # Estilo para encabezados
        for col in ['A', 'B', 'C']:
            cell = chart_ws[f'{col}1']
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        row = 2
        total = stats['total_analyses']
        for category, count in stats['quality_distribution'].items():
            percentage = (count / total * 100) if total > 0 else 0
            chart_ws[f'A{row}'] = category
            chart_ws[f'B{row}'] = count
            chart_ws[f'C{row}'] = f"{percentage:.1f}%"
            row += 1
        
        # Crear gráfico de barras
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Distribución de Calidad"
        chart.y_axis.title = 'Cantidad'
        chart.x_axis.title = 'Categoría'
        
        data = Reference(chart_ws, min_col=2, min_row=1, max_row=row-1, max_col=2)
        cats = Reference(chart_ws, min_col=1, min_row=2, max_row=row-1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        chart_ws.add_chart(chart, "E2")
        
        # Ajustar ancho de columnas
        chart_ws.column_dimensions['A'].width = 20
        chart_ws.column_dimensions['B'].width = 12
        chart_ws.column_dimensions['C'].width = 12
    
    def _create_summary_sheet(self, stats, user):
        """Crear hoja de resumen."""
        summary_ws = self.workbook.create_sheet("Resumen")
        
        # Título
        summary_ws['A1'] = "Resumen Ejecutivo"
        summary_ws['A1'].font = Font(size=16, bold=True, color="2F4F4F")
        summary_ws['A1'].alignment = Alignment(horizontal='center')
        summary_ws.merge_cells('A1:D1')
        
        # Información del reporte
        summary_ws['A3'] = f"Generado el: {timezone.now().strftime('%d/%m/%Y %H:%M')}"
        summary_ws['A3'].font = Font(size=10, italic=True)
        
        summary_ws['A4'] = f"Usuario: {user.get_full_name() or user.username}"
        summary_ws['A4'].font = Font(size=10, italic=True)
        
        # Resumen de métricas clave
        summary_ws['A6'] = "Métricas Clave"
        summary_ws['A6'].font = Font(size=14, bold=True, color="2F4F4F")
        
        key_metrics = [
            ['Total de Análisis', stats['total_analyses']],
            ['Confianza Promedio', f"{stats['avg_confidence']}%"],
            ['Alto Promedio', f"{stats['avg_dimensions']['alto']} mm"],
            ['Ancho Promedio', f"{stats['avg_dimensions']['ancho']} mm"],
            ['Grosor Promedio', f"{stats['avg_dimensions']['grosor']} mm"],
            ['Peso Promedio', f"{stats['avg_weight']} g"],
        ]
        
        for row_num, (metric, value) in enumerate(key_metrics, 8):
            summary_ws[f'A{row_num}'] = metric
            summary_ws[f'B{row_num}'] = value
            summary_ws[f'A{row_num}'].font = Font(bold=True)
            summary_ws[f'B{row_num}'].alignment = Alignment(horizontal='center')
        
        # Recomendaciones
        summary_ws['A15'] = "Recomendaciones"
        summary_ws['A15'].font = Font(size=14, bold=True, color="2F4F4F")
        
        recommendations = []
        if stats['avg_confidence'] < 70:
            recommendations.append("• La confianza promedio es baja. Considere mejorar la calidad de las imágenes.")
        if stats['avg_dimensions']['alto'] < 15 or stats['avg_dimensions']['alto'] > 25:
            recommendations.append("• Las dimensiones están fuera del rango óptimo. Revise el proceso de cosecha.")
        if stats['avg_weight'] < 1.0 or stats['avg_weight'] > 2.5:
            recommendations.append("• El peso promedio no está en el rango esperado. Verifique la madurez.")
        
        if not recommendations:
            recommendations.append("• Los indicadores están dentro de rangos aceptables. Mantenga las buenas prácticas.")
        
        for row_num, rec in enumerate(recommendations, 17):
            summary_ws[f'A{row_num}'] = rec
            summary_ws[f'A{row_num}'].font = Font(size=10)
        
        # Ajustar ancho de columnas
        summary_ws.column_dimensions['A'].width = 25
        summary_ws.column_dimensions['B'].width = 15
    
    def _create_finca_info_section(self, finca):
        """Crear sección de información de finca."""
        # Título de sección
        self.ws['A8'] = "Información de la Finca"
        self.ws['A8'].font = Font(size=14, bold=True, color="2F4F4F")
        
        # Datos de la finca
        finca_data = [
            ['Campo', 'Valor'],
            ['Nombre', finca.nombre],
            ['Ubicación', finca.ubicacion_completa],
            ['Hectáreas', f"{finca.hectareas} ha"],
            ['Agricultor', finca.agricultor.get_full_name() or finca.agricultor.username],
            ['Total de Lotes', str(finca.total_lotes)],
            ['Lotes Activos', str(finca.lotes_activos)],
            ['Total de Análisis', str(finca.total_analisis)],
            ['Calidad Promedio', f"{finca.calidad_promedio}%"],
        ]
        
        # Crear tabla
        for row_num, row_data in enumerate(finca_data, 10):
            for col_num, cell_value in enumerate(row_data, 1):
                cell = self.ws.cell(row=row_num, column=col_num, value=cell_value)
                
                # Estilo para encabezados
                if row_num == 10:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center')
                else:
                    cell.alignment = Alignment(horizontal='left')
                
                # Bordes
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # Ajustar ancho de columnas
        self.ws.column_dimensions['A'].width = 20
        self.ws.column_dimensions['B'].width = 30
    
    def _get_lotes_stats(self, finca):
        """Obtener estadísticas de lotes de la finca."""
        lotes = finca.lotes.all()
        
        return {
            'total_lotes': lotes.count(),
            'lotes_activos': lotes.filter(activo=True).count(),
            'total_area': sum(float(lote.area_hectareas) for lote in lotes),
            'variedades': list(lotes.values('variedad').distinct()),
            'estados': dict(lotes.values('estado').annotate(count=Count('id')).values_list('estado', 'count')),
        }
    
    def _create_lotes_stats_section(self, stats):
        """Crear sección de estadísticas de lotes."""
        # Título de sección
        self.ws['A20'] = "Estadísticas de Lotes"
        self.ws['A20'].font = Font(size=14, bold=True, color="2F4F4F")
        
        # Datos de estadísticas
        data = [
            ['Métrica', 'Valor'],
            ['Total de Lotes', str(stats['total_lotes'])],
            ['Lotes Activos', str(stats['lotes_activos'])],
            ['Área Total', f"{stats['total_area']:.2f} ha"],
            ['Variedades', str(len(stats['variedades']))],
        ]
        
        # Crear tabla
        for row_num, row_data in enumerate(data, 22):
            for col_num, cell_value in enumerate(row_data, 1):
                cell = self.ws.cell(row=row_num, column=col_num, value=cell_value)
                
                # Estilo para encabezados
                if row_num == 22:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center')
                else:
                    cell.alignment = Alignment(horizontal='center')
                
                # Bordes
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # Ajustar ancho de columnas
        self.ws.column_dimensions['A'].width = 20
        self.ws.column_dimensions['B'].width = 15
    
    def _create_lotes_analysis_section(self, finca):
        """Crear sección de análisis por lote."""
        # Título de sección
        self.ws['A28'] = "Análisis por Lote"
        self.ws['A28'].font = Font(size=14, bold=True, color="2F4F4F")
        
        lotes = finca.lotes.all()
        
        if lotes.exists():
            # Encabezados
            headers = ['Lote', 'Variedad', 'Estado', 'Área (ha)', 'Análisis', 'Calidad (%)']
            for col_num, header in enumerate(headers, 1):
                cell = self.ws.cell(row=30, column=col_num, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            
            # Datos de lotes
            for row_num, lote in enumerate(lotes, 31):
                data = [
                    lote.identificador,
                    lote.variedad,
                    lote.get_estado_display(),
                    f"{lote.area_hectareas:.2f}",
                    str(lote.total_analisis),
                    f"{lote.calidad_promedio:.1f}",
                ]
                
                for col_num, value in enumerate(data, 1):
                    cell = self.ws.cell(row=row_num, column=col_num, value=value)
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
            
            # Ajustar ancho de columnas
            column_widths = [12, 15, 12, 12, 10, 12]
            for i, width in enumerate(column_widths, 1):
                self.ws.column_dimensions[chr(64 + i)].width = width
    
    def _create_detailed_lotes_sheet(self, finca):
        """Crear hoja detallada de lotes."""
        lotes_ws = self.workbook.create_sheet("Lotes Detallados")
        
        # Título
        lotes_ws['A1'] = f"Análisis Detallados - Finca {finca.nombre}"
        lotes_ws['A1'].font = Font(size=16, bold=True, color="2F4F4F")
        lotes_ws['A1'].alignment = Alignment(horizontal='center')
        lotes_ws.merge_cells('A1:H1')
        
        # Encabezados
        headers = ['Lote', 'Variedad', 'Estado', 'Área (ha)', 'Fecha Plantación', 'Análisis', 'Calidad (%)', 'Observaciones']
        for col_num, header in enumerate(headers, 1):
            cell = lotes_ws.cell(row=3, column=col_num, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Datos de lotes
        lotes = finca.lotes.all()
        for row_num, lote in enumerate(lotes, 4):
            data = [
                lote.identificador,
                lote.variedad,
                lote.get_estado_display(),
                f"{lote.area_hectareas:.2f}",
                lote.fecha_plantacion.strftime('%d/%m/%Y'),
                str(lote.total_analisis),
                f"{lote.calidad_promedio:.1f}",
                lote.descripcion or 'Sin observaciones',
            ]
            
            for col_num, value in enumerate(data, 1):
                cell = lotes_ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(horizontal='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # Ajustar ancho de columnas
        column_widths = [12, 15, 12, 12, 15, 10, 12, 25]
        for i, width in enumerate(column_widths, 1):
            lotes_ws.column_dimensions[chr(64 + i)].width = width
    
    def _get_activity_stats(self, filtros):
        """Obtener estadísticas de actividad."""
        queryset = ActivityLog.objects.all()
        
        if filtros:
            if filtros.get('fecha_desde'):
                queryset = queryset.filter(timestamp__date__gte=filtros['fecha_desde'])
            if filtros.get('fecha_hasta'):
                queryset = queryset.filter(timestamp__date__lte=filtros['fecha_hasta'])
        
        return {
            'total_activities': queryset.count(),
            'activities_today': queryset.filter(timestamp__date=timezone.now().date()).count(),
            'activities_by_action': dict(queryset.values('accion').annotate(count=Count('id')).values_list('accion', 'count')),
            'top_users': list(queryset.values('usuario__username').annotate(count=Count('id')).order_by('-count')[:10]),
        }
    
    def _create_activity_stats_section(self, stats):
        """Crear sección de estadísticas de actividad."""
        # Título de sección
        self.ws['A8'] = "Estadísticas de Actividad"
        self.ws['A8'].font = Font(size=14, bold=True, color="2F4F4F")
        
        # Datos de estadísticas
        data = [
            ['Métrica', 'Valor'],
            ['Total de Actividades', str(stats['total_activities'])],
            ['Actividades Hoy', str(stats['activities_today'])],
        ]
        
        # Crear tabla
        for row_num, row_data in enumerate(data, 10):
            for col_num, cell_value in enumerate(row_data, 1):
                cell = self.ws.cell(row=row_num, column=col_num, value=cell_value)
                
                # Estilo para encabezados
                if row_num == 10:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center')
                else:
                    cell.alignment = Alignment(horizontal='center')
                
                # Bordes
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # Ajustar ancho de columnas
        self.ws.column_dimensions['A'].width = 20
        self.ws.column_dimensions['B'].width = 15
    
    def _get_login_stats(self, filtros):
        """Obtener estadísticas de logins."""
        queryset = LoginHistory.objects.all()
        
        if filtros:
            if filtros.get('fecha_desde'):
                queryset = queryset.filter(login_time__date__gte=filtros['fecha_desde'])
            if filtros.get('fecha_hasta'):
                queryset = queryset.filter(login_time__date__lte=filtros['fecha_hasta'])
        
        return {
            'total_logins': queryset.count(),
            'successful_logins': queryset.filter(success=True).count(),
            'failed_logins': queryset.filter(success=False).count(),
            'success_rate': (queryset.filter(success=True).count() / queryset.count() * 100) if queryset.count() > 0 else 0,
        }
    
    def _create_login_stats_section(self, stats):
        """Crear sección de estadísticas de logins."""
        # Título de sección
        self.ws['A14'] = "Estadísticas de Logins"
        self.ws['A14'].font = Font(size=14, bold=True, color="2F4F4F")
        
        # Datos de estadísticas
        data = [
            ['Métrica', 'Valor'],
            ['Total de Logins', str(stats['total_logins'])],
            ['Logins Exitosos', str(stats['successful_logins'])],
            ['Logins Fallidos', str(stats['failed_logins'])],
            ['Tasa de Éxito', f"{stats['success_rate']:.1f}%"],
        ]
        
        # Crear tabla
        for row_num, row_data in enumerate(data, 16):
            for col_num, cell_value in enumerate(row_data, 1):
                cell = self.ws.cell(row=row_num, column=col_num, value=cell_value)
                
                # Estilo para encabezados
                if row_num == 16:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center')
                else:
                    cell.alignment = Alignment(horizontal='center')
                
                # Bordes
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # Ajustar ancho de columnas
        self.ws.column_dimensions['A'].width = 20
        self.ws.column_dimensions['B'].width = 15
    
    def _create_detailed_activities_sheet(self, filtros):
        """Crear hoja detallada de actividades."""
        activities_ws = self.workbook.create_sheet("Actividades Detalladas")
        
        # Título
        activities_ws['A1'] = "Actividades del Sistema"
        activities_ws['A1'].font = Font(size=16, bold=True, color="2F4F4F")
        activities_ws['A1'].alignment = Alignment(horizontal='center')
        activities_ws.merge_cells('A1:F1')
        
        # Encabezados
        headers = ['Fecha', 'Usuario', 'Acción', 'Modelo', 'Descripción', 'IP']
        for col_num, header in enumerate(headers, 1):
            cell = activities_ws.cell(row=3, column=col_num, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Datos de actividades (limitar a 100 registros)
        queryset = ActivityLog.objects.select_related('usuario').order_by('-timestamp')[:100]
        
        if filtros:
            if filtros.get('fecha_desde'):
                queryset = queryset.filter(timestamp__date__gte=filtros['fecha_desde'])
            if filtros.get('fecha_hasta'):
                queryset = queryset.filter(timestamp__date__lte=filtros['fecha_hasta'])
        
        for row_num, activity in enumerate(queryset, 4):
            data = [
                activity.timestamp.strftime('%d/%m/%Y %H:%M'),
                activity.usuario.username if activity.usuario else 'Anónimo',
                activity.get_accion_display(),
                activity.modelo,
                activity.descripcion[:50] + '...' if len(activity.descripcion) > 50 else activity.descripcion,
                activity.ip_address or 'N/A',
            ]
            
            for col_num, value in enumerate(data, 1):
                cell = activities_ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(horizontal='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # Ajustar ancho de columnas
        column_widths = [18, 12, 12, 12, 30, 15]
        for i, width in enumerate(column_widths, 1):
            activities_ws.column_dimensions[chr(64 + i)].width = width
    
    def _create_detailed_logins_sheet(self, filtros):
        """Crear hoja detallada de logins."""
        logins_ws = self.workbook.create_sheet("Logins Detallados")
        
        # Título
        logins_ws['A1'] = "Historial de Logins"
        logins_ws['A1'].font = Font(size=16, bold=True, color="2F4F4F")
        logins_ws['A1'].alignment = Alignment(horizontal='center')
        logins_ws.merge_cells('A1:F1')
        
        # Encabezados
        headers = ['Fecha', 'Usuario', 'IP', 'Éxito', 'Duración', 'Razón Fallo']
        for col_num, header in enumerate(headers, 1):
            cell = logins_ws.cell(row=3, column=col_num, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Datos de logins (limitar a 100 registros)
        queryset = LoginHistory.objects.select_related('usuario').order_by('-login_time')[:100]
        
        if filtros:
            if filtros.get('fecha_desde'):
                queryset = queryset.filter(login_time__date__gte=filtros['fecha_desde'])
            if filtros.get('fecha_hasta'):
                queryset = queryset.filter(login_time__date__lte=filtros['fecha_hasta'])
        
        for row_num, login in enumerate(queryset, 4):
            duration = login.session_duration_formatted if hasattr(login, 'session_duration_formatted') else 'N/A'
            data = [
                login.login_time.strftime('%d/%m/%Y %H:%M'),
                login.usuario.username,
                login.ip_address,
                'Sí' if login.success else 'No',
                duration,
                login.failure_reason or 'N/A',
            ]
            
            for col_num, value in enumerate(data, 1):
                cell = logins_ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(horizontal='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # Ajustar ancho de columnas
        column_widths = [18, 12, 15, 8, 12, 20]
        for i, width in enumerate(column_widths, 1):
            logins_ws.column_dimensions[chr(64 + i)].width = width
    
    def _create_custom_quality_section(self, stats, parametros):
        """Crear sección personalizada de calidad."""
        # Título de sección
        self.ws['A8'] = "Análisis Personalizado de Calidad"
        self.ws['A8'].font = Font(size=14, bold=True, color="2F4F4F")
        
        # Métricas específicas según parámetros
        custom_metrics = []
        
        if parametros.get('include_dimensions', True):
            custom_metrics.extend([
                ['Alto Promedio', f"{stats['avg_dimensions']['alto']} mm"],
                ['Ancho Promedio', f"{stats['avg_dimensions']['ancho']} mm"],
                ['Grosor Promedio', f"{stats['avg_dimensions']['grosor']} mm"],
            ])
        
        if parametros.get('include_weight', True):
            custom_metrics.append(['Peso Promedio', f"{stats['avg_weight']} g"])
        
        if parametros.get('include_confidence', True):
            custom_metrics.append(['Confianza Promedio', f"{stats['avg_confidence']}%"])
        
        # Crear tabla personalizada
        data = [['Métrica', 'Valor']] + custom_metrics
        
        for row_num, row_data in enumerate(data, 10):
            for col_num, cell_value in enumerate(row_data, 1):
                cell = self.ws.cell(row=row_num, column=col_num, value=cell_value)
                
                # Estilo para encabezados
                if row_num == 10:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center')
                else:
                    cell.alignment = Alignment(horizontal='center')
                
                # Bordes
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # Ajustar ancho de columnas
        self.ws.column_dimensions['A'].width = 20
        self.ws.column_dimensions['B'].width = 15
    
    def _create_custom_finca_section(self, finca, parametros):
        """Crear sección personalizada de finca."""
        # Título de sección
        self.ws['A8'] = f"Análisis Personalizado - {finca.nombre}"
        self.ws['A8'].font = Font(size=14, bold=True, color="2F4F4F")
        
        # Métricas específicas según parámetros
        custom_metrics = []
        
        if parametros.get('include_basic_info', True):
            custom_metrics.extend([
                ['Nombre', finca.nombre],
                ['Ubicación', finca.ubicacion_completa],
                ['Hectáreas', f"{finca.hectareas} ha"],
            ])
        
        if parametros.get('include_lotes', True):
            custom_metrics.extend([
                ['Total de Lotes', str(finca.total_lotes)],
                ['Lotes Activos', str(finca.lotes_activos)],
            ])
        
        if parametros.get('include_quality', True):
            custom_metrics.extend([
                ['Total de Análisis', str(finca.total_analisis)],
                ['Calidad Promedio', f"{finca.calidad_promedio}%"],
            ])
        
        # Crear tabla personalizada
        data = [['Campo', 'Valor']] + custom_metrics
        
        for row_num, row_data in enumerate(data, 10):
            for col_num, cell_value in enumerate(row_data, 1):
                cell = self.ws.cell(row=row_num, column=col_num, value=cell_value)
                
                # Estilo para encabezados
                if row_num == 10:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center')
                else:
                    cell.alignment = Alignment(horizontal='left')
                
                # Bordes
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # Ajustar ancho de columnas
        self.ws.column_dimensions['A'].width = 20
        self.ws.column_dimensions['B'].width = 30
    
    def _create_custom_audit_section(self, stats, parametros):
        """Crear sección personalizada de auditoría."""
        # Título de sección
        self.ws['A8'] = "Análisis Personalizado de Auditoría"
        self.ws['A8'].font = Font(size=14, bold=True, color="2F4F4F")
        
        # Métricas específicas según parámetros
        custom_metrics = []
        
        if parametros.get('include_activity', True):
            custom_metrics.extend([
                ['Total de Actividades', str(stats['total_activities'])],
                ['Actividades Hoy', str(stats['activities_today'])],
            ])
        
        if parametros.get('include_top_users', True):
            custom_metrics.append(['Usuarios Más Activos', str(len(stats['top_users']))])
        
        if parametros.get('include_action_types', True):
            custom_metrics.append(['Tipos de Acción', str(len(stats['activities_by_action']))])
        
        # Crear tabla personalizada
        data = [['Métrica', 'Valor']] + custom_metrics
        
        for row_num, row_data in enumerate(data, 10):
            for col_num, cell_value in enumerate(row_data, 1):
                cell = self.ws.cell(row=row_num, column=col_num, value=cell_value)
                
                # Estilo para encabezados
                if row_num == 10:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center')
                else:
                    cell.alignment = Alignment(horizontal='center')
                
                # Bordes
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # Ajustar ancho de columnas
        self.ws.column_dimensions['A'].width = 20
        self.ws.column_dimensions['B'].width = 15