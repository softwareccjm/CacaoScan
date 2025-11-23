"""
Vistas para gestión de reportes avanzados en CacaoScan.
"""
import logging
import io
import json
import traceback
from datetime import datetime, timedelta
from django.utils import timezone
from django.core.files.base import ContentFile
from django.http import HttpResponse, FileResponse
from django.contrib.auth import get_user_model
from rest_framework.permissions import IsAuthenticated, IsAdminUser
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.renderers import BaseRenderer
from rest_framework import status
from django.core.paginator import Paginator
from django.db.models import Count  # Importar Count
from drf_yasg.utils import swagger_auto_schema
from drf_yasg import openapi

from ..mixins import PaginationMixin, AdminPermissionMixin

from reports.models import ReporteGenerado
from ...utils.model_imports import get_models_safely

# Import models safely
models = get_models_safely({
    'CacaoImage': 'images_app.models.CacaoImage',
    'CacaoPrediction': 'images_app.models.CacaoPrediction',
    'Finca': 'fincas_app.models.Finca',
    'Lote': 'fincas_app.models.Lote'
})
CacaoImage = models['CacaoImage']
CacaoPrediction = models['CacaoPrediction']
Finca = models['Finca']
Lote = models['Lote']
from ...services.report import CacaoReportPDFGenerator
from ...services.report.excel import ExcelAgricultoresService, ExcelUsuariosService, ExcelAnalisisService
from ...serializers import ErrorResponseSerializer

logger = logging.getLogger("cacaoscan.api")


class ExcelRenderer(BaseRenderer):
    """
    Renderer personalizado para archivos Excel binarios.
    Evita que DRF intente serializar el contenido binario como JSON.
    """
    media_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    format = 'xlsx'
    charset = None
    
    def render(self, data, accepted_media_type=None, renderer_context=None):
        """
        Renderiza los datos binarios directamente sin procesamiento.
        Si data es HttpResponse, retornamos el contenido tal cual.
        Si data es bytes, lo devolvemos directamente.
        """
        if isinstance(data, HttpResponse):
            # Si ya es HttpResponse, extraer el contenido binario
            return data.content
        if isinstance(data, bytes):
            # Si es bytes directamente, devolverlo
            return data
        # Por defecto, retornar los datos tal cual
        return data


class ReporteListCreateView(PaginationMixin, APIView):
    """
    Vista para listar y crear reportes.
    GET: Lista reportes del usuario
    POST: Crea nuevo reporte
    """
    permission_classes = [IsAuthenticated]
    
    @swagger_auto_schema(
        operation_description="Lista todos los reportes del usuario autenticado",
        operation_summary="Listar reportes",
        manual_parameters=[
            openapi.Parameter('tipo_reporte', openapi.IN_QUERY, description="Filtrar por tipo de reporte", type=openapi.TYPE_STRING),
            openapi.Parameter('formato', openapi.IN_QUERY, description="Filtrar por formato", type=openapi.TYPE_STRING),
            openapi.Parameter('estado', openapi.IN_QUERY, description="Filtrar por estado", type=openapi.TYPE_STRING),
            openapi.Parameter('page', openapi.IN_QUERY, description="Número de página", type=openapi.TYPE_INTEGER),
            openapi.Parameter('page_size', openapi.IN_QUERY, description="Tamaño de página", type=openapi.TYPE_INTEGER),
        ],
        responses={
            200: openapi.Response(description="Lista de reportes obtenida exitosamente"),
            401: ErrorResponseSerializer,
        },
        tags=['Reportes']
    )
    def get(self, request):
        """Listar reportes con filtros."""
        try:
            queryset = ReporteGenerado.objects.filter(usuario=request.user)
            
            # Aplicar filtros
            tipo_reporte = request.GET.get('tipo_reporte', '').strip()
            if tipo_reporte:
                queryset = queryset.filter(tipo_reporte=tipo_reporte)
            
            formato = request.GET.get('formato', '').strip()
            if formato:
                queryset = queryset.filter(formato=formato)
            
            estado = request.GET.get('estado', '').strip()
            if estado:
                queryset = queryset.filter(estado=estado)
            
            # Función de serialización personalizada
            def serialize_reportes(reportes):
                reportes_data = []
                for reporte in reportes:
                    reportes_data.append({
                        'id': reporte.id,
                        'tipo_reporte': reporte.tipo_reporte,
                        'tipo_reporte_display': reporte.get_tipo_reporte_display(),
                        'formato': reporte.formato,
                        'formato_display': reporte.get_formato_display(),
                        'titulo': reporte.titulo,
                        'descripcion': reporte.descripcion,
                        'estado': reporte.estado,
                        'estado_display': reporte.get_estado_display(),
                        'fecha_solicitud': reporte.fecha_solicitud.isoformat(),
                        'fecha_generacion': reporte.fecha_generacion.isoformat() if reporte.fecha_generacion else None,
                        'fecha_expiracion': reporte.fecha_expiracion.isoformat() if reporte.fecha_expiracion else None,
                        'tiempo_generacion_segundos': reporte.tiempo_generacion_segundos,
                        'tamano_archivo_mb': reporte.tamano_archivo_mb,
                        'archivo_url': reporte.archivo_url,
                        'esta_expirado': reporte.esta_expirado,
                        'mensaje_error': reporte.mensaje_error,
                    })
                return reportes_data
            
            # Paginar usando el mixin con serialización personalizada
            return self.paginate_queryset(
                request,
                queryset,
                serializer_func=serialize_reportes
            )
            
        except Exception as e:
            logger.error(f"Error listando reportes: {e}")
            return Response({
                'error': 'Error interno del servidor',
                'status': 'error'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @swagger_auto_schema(
        operation_description="Crea un nuevo reporte",
        operation_summary="Crear reporte",
        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            properties={
                'tipo_reporte': openapi.Schema(type=openapi.TYPE_STRING, description="Tipo de reporte"),
                'formato': openapi.Schema(type=openapi.TYPE_STRING, description="Formato del reporte"),
                'titulo': openapi.Schema(type=openapi.TYPE_STRING, description="Título del reporte"),
                'descripcion': openapi.Schema(type=openapi.TYPE_STRING, description="Descripción del reporte"),
                'parametros': openapi.Schema(type=openapi.TYPE_OBJECT, description="Parámetros del reporte"),
                'filtros': openapi.Schema(type=openapi.TYPE_OBJECT, description="Filtros a aplicar"),
            },
            required=['tipo_reporte', 'formato', 'titulo']
        ),
        responses={
            201: openapi.Response(description="Reporte creado exitosamente"),
            400: ErrorResponseSerializer,
            401: ErrorResponseSerializer,
        },
        tags=['Reportes']
    )
    def post(self, request):
        """Crear nuevo reporte."""
        try:
            # Validar datos de entrada
            tipo_reporte = request.data.get('tipo_reporte')
            formato = request.data.get('formato')
            titulo = request.data.get('titulo')
            
            if not tipo_reporte or not formato or not titulo:
                return Response({
                    'error': 'Los campos tipo_reporte, formato y titulo son requeridos',
                    'status': 'error'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Validar tipo de reporte
            valid_types = [choice[0] for choice in ReporteGenerado.TIPO_REPORTE_CHOICES]
            if tipo_reporte not in valid_types:
                return Response({
                    'error': f'Tipo de reporte inválido. Opciones válidas: {", ".join(valid_types)}',
                    'status': 'error'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Validar formato
            valid_formats = [choice[0] for choice in ReporteGenerado.FORMATO_CHOICES]
            if formato not in valid_formats:
                return Response({
                    'error': f'Formato inválido. Opciones válidas: {", ".join(valid_formats)}',
                    'status': 'error'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Crear reporte
            reporte = ReporteGenerado.generar_reporte(
                usuario=request.user,
                tipo_reporte=tipo_reporte,
                formato=formato,
                titulo=titulo,
                descripcion=request.data.get('descripcion'),
                parametros=request.data.get('parametros', {}),
                filtros=request.data.get('filtros', {})
            )
            
            # Generar reporte en background (simulado)
            self._generate_report_async(reporte)
            
            logger.info(f"Reporte '{titulo}' creado por usuario {request.user.username}")
            
            return Response({
                'id': reporte.id,
                'tipo_reporte': reporte.tipo_reporte,
                'formato': reporte.formato,
                'titulo': reporte.titulo,
                'estado': reporte.estado,
                'fecha_solicitud': reporte.fecha_solicitud.isoformat(),
                'message': 'Reporte creado exitosamente. Se generará en segundo plano.',
                'status': 'success'
            }, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            logger.error(f"Error creando reporte: {e}")
            return Response({
                'error': 'Error interno del servidor',
                'status': 'error'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def _generate_report_async(self, reporte):
        """Generar reporte de forma asíncrona (simulado)."""
        try:
            start_time = timezone.now()
            user = reporte.usuario  # Get user from reporte object
            
            # Generar según tipo y formato
            if reporte.formato == 'pdf':
                generator = CacaoReportPDFGenerator()
                # PDF generation logic (to be implemented if needed)
                raise ValueError("Generación de PDF no implementada aún")
            elif reporte.formato == 'excel':
                excel_service = ExcelAnalisisService()
                # Generar contenido según tipo
                if reporte.tipo_reporte == 'calidad':
                    content = excel_service.generate_quality_report(user, reporte.filtros_aplicados)
                elif reporte.tipo_reporte == 'finca':
                    finca_id = reporte.parametros.get('finca_id')
                    if not finca_id:
                        raise ValueError("finca_id es requerido para reportes de finca")
                    content = excel_service.generate_finca_report(finca_id, user, reporte.filtros_aplicados)
                elif reporte.tipo_reporte == 'auditoria':
                    content = excel_service.generate_audit_report(user, reporte.filtros_aplicados)
                elif reporte.tipo_reporte == 'personalizado':
                    content = excel_service.generate_custom_report(
                        user, 
                        reporte.parametros.get('tipo_reporte', 'calidad'),
                        reporte.parametros,
                        reporte.filtros_aplicados
                    )
                else:
                    raise ValueError(f"Tipo de reporte no soportado: {reporte.tipo_reporte}")
            else:
                raise ValueError(f"Formato no soportado: {reporte.formato}")
            
            # Crear archivo
            filename = f"{reporte.titulo}_{reporte.id}.{reporte.formato}"
            file_content = ContentFile(content)
            
            # Calcular tiempo de generación
            end_time = timezone.now()
            tiempo_generacion = end_time - start_time
            
            # Marcar como completado
            reporte.marcar_completado(file_content, tiempo_generacion)
            
            logger.info(f"Reporte {reporte.id} generado exitosamente en {tiempo_generacion.total_seconds():.2f} segundos")
            
        except Exception as e:
            logger.error(f"Error generando reporte {reporte.id}: {e}")
            reporte.marcar_fallido(str(e))


class ReporteDetailView(APIView):
    """
    Vista para obtener detalles de un reporte específico.
    """
    permission_classes = [IsAuthenticated]
    
    @swagger_auto_schema(
        operation_description="Obtiene los detalles de un reporte específico",
        operation_summary="Detalles de reporte",
        responses={
            200: openapi.Response(description="Detalles de reporte obtenidos exitosamente"),
            404: ErrorResponseSerializer,
            401: ErrorResponseSerializer,
        },
        tags=['Reportes']
    )
    def get(self, request, reporte_id):
        """Obtener detalles de reporte."""
        try:
            reporte = ReporteGenerado.objects.get(id=reporte_id, usuario=request.user)
            
            return Response({
                'id': reporte.id,
                'tipo_reporte': reporte.tipo_reporte,
                'tipo_reporte_display': reporte.get_tipo_reporte_display(),
                'formato': reporte.formato,
                'formato_display': reporte.get_formato_display(),
                'titulo': reporte.titulo,
                'descripcion': reporte.descripcion,
                'estado': reporte.estado,
                'estado_display': reporte.get_estado_display(),
                'fecha_solicitud': reporte.fecha_solicitud.isoformat(),
                'fecha_generacion': reporte.fecha_generacion.isoformat() if reporte.fecha_generacion else None,
                'fecha_expiracion': reporte.fecha_expiracion.isoformat() if reporte.fecha_expiracion else None,
                'tiempo_generacion_segundos': reporte.tiempo_generacion_segundos,
                'tamano_archivo_mb': reporte.tamano_archivo_mb,
                'archivo_url': reporte.archivo_url,
                'esta_expirado': reporte.esta_expirado,
                'mensaje_error': reporte.mensaje_error,
                'parametros': reporte.parametros,
                'filtros_aplicados': reporte.filtros_aplicados,
            }, status=status.HTTP_200_OK)
            
        except ReporteGenerado.DoesNotExist:
            return Response({
                'error': 'Reporte no encontrado',
                'status': 'error'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"Error obteniendo detalles de reporte {reporte_id}: {e}")
            return Response({
                'error': 'Error interno del servidor',
                'status': 'error'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class ReporteDownloadView(APIView):
    """
    Vista para descargar un reporte generado.
    """
    permission_classes = [IsAuthenticated]
    
    @swagger_auto_schema(
        operation_description="Descarga un reporte generado",
        operation_summary="Descargar reporte",
        responses={
            200: openapi.Response(description="Archivo descargado exitosamente"),
            404: ErrorResponseSerializer,
            410: ErrorResponseSerializer,
            401: ErrorResponseSerializer,
        },
        tags=['Reportes']
    )
    def get(self, request, reporte_id):
        """Descargar reporte."""
        try:
            reporte = ReporteGenerado.objects.get(id=reporte_id, usuario=request.user)
            
            # Verificar estado
            if reporte.estado != 'completado':
                return Response({
                    'error': 'El reporte aún no está listo para descarga',
                    'status': 'error'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Verificar si está expirado
            if reporte.esta_expirado:
                return Response({
                    'error': 'El reporte ha expirado y ya no está disponible',
                    'status': 'error'
                }, status=status.HTTP_410_GONE)
            
            # Verificar que existe el archivo
            if not reporte.archivo:
                return Response({
                    'error': 'El archivo del reporte no está disponible',
                    'status': 'error'
                }, status=status.HTTP_404_NOT_FOUND)
            
            # Preparar respuesta de descarga
            response = FileResponse(
                reporte.archivo,
                as_attachment=True,
                filename=reporte.nombre_archivo or f"{reporte.titulo}.{reporte.formato}"
            )
            
            # Configurar headers según formato
            if reporte.formato == 'pdf':
                response['Content-Type'] = 'application/pdf'
            elif reporte.formato == 'excel':
                response['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            elif reporte.formato == 'csv':
                response['Content-Type'] = 'text/csv'
            elif reporte.formato == 'json':
                response['Content-Type'] = 'application/json'
            
            logger.info(f"Reporte {reporte_id} descargado por usuario {request.user.username}")
            
            return response
            
        except ReporteGenerado.DoesNotExist:
            return Response({
                'error': 'Reporte no encontrado',
                'status': 'error'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"Error descargando reporte {reporte_id}: {e}")
            return Response({
                'error': 'Error interno del servidor',
                'status': 'error'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class ReporteDeleteView(APIView):
    """
    Vista para eliminar un reporte.
    """
    permission_classes = [IsAuthenticated]
    
    @swagger_auto_schema(
        operation_description="Elimina un reporte específico",
        operation_summary="Eliminar reporte",
        responses={
            204: "Reporte eliminado exitosamente",
            404: ErrorResponseSerializer,
            401: ErrorResponseSerializer,
        },
        tags=['Reportes']
    )
    def delete(self, request, reporte_id):
        """Eliminar reporte."""
        try:
            reporte = ReporteGenerado.objects.get(id=reporte_id, usuario=request.user)
            
            # Eliminar archivo físico si existe
            if reporte.archivo:
                try:
                    reporte.archivo.delete(save=False)
                except Exception as e:
                    logger.warning(f"No se pudo eliminar archivo físico del reporte {reporte_id}: {e}")
            
            # Eliminar registro
            reporte.delete()
            
            logger.info(f"Reporte {reporte_id} eliminado por usuario {request.user.username}")
            
            return Response(status=status.HTTP_204_NO_CONTENT)
            
        except ReporteGenerado.DoesNotExist:
            return Response({
                'error': 'Reporte no encontrado',
                'status': 'error'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"Error eliminando reporte {reporte_id}: {e}")
            return Response({
                'error': 'Error interno del servidor',
                'status': 'error'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class ReporteStatsView(APIView):
    """
    Vista para obtener estadísticas de reportes del usuario.
    """
    permission_classes = [IsAuthenticated]
    
    @swagger_auto_schema(
        operation_description="Obtiene estadísticas de reportes del usuario",
        operation_summary="Estadísticas de reportes",
        responses={
            200: openapi.Response(description="Estadísticas obtenidas exitosamente"),
            401: ErrorResponseSerializer,
        },
        tags=['Reportes']
    )
    def get(self, request):
        """Obtener estadísticas de reportes."""
        try:
            # Estadísticas básicas
            total_reportes = ReporteGenerado.objects.filter(usuario=request.user).count()
            reportes_completados = ReporteGenerado.objects.filter(usuario=request.user, estado='completado').count()
            reportes_generando = ReporteGenerado.objects.filter(usuario=request.user, estado='generando').count()
            reportes_fallidos = ReporteGenerado.objects.filter(usuario=request.user, estado='fallido').count()
            
            # Reportes por tipo
            reportes_por_tipo = dict(
                ReporteGenerado.objects.filter(usuario=request.user)
                .values('tipo_reporte')
                .annotate(count=Count('id'))
                .values_list('tipo_reporte', 'count')
            )
            
            # Reportes por formato
            reportes_por_formato = dict(
                ReporteGenerado.objects.filter(usuario=request.user)
                .values('formato')
                .annotate(count=Count('id'))
                .values_list('formato', 'count')
            )
            
            # Reportes recientes (últimos 5)
            reportes_recientes = ReporteGenerado.objects.filter(usuario=request.user).order_by('-fecha_solicitud')[:5]
            reportes_recientes_data = []
            for reporte in reportes_recientes:
                reportes_recientes_data.append({
                    'id': reporte.id,
                    'titulo': reporte.titulo,
                    'tipo_reporte': reporte.tipo_reporte,
                    'formato': reporte.formato,
                    'estado': reporte.estado,
                    'fecha_solicitud': reporte.fecha_solicitud.isoformat(),
                })
            
            stats = {
                'total_reportes': total_reportes,
                'reportes_completados': reportes_completados,
                'reportes_generando': reportes_generando,
                'reportes_fallidos': reportes_fallidos,
                'reportes_por_tipo': reportes_por_tipo,
                'reportes_por_formato': reportes_por_formato,
                'reportes_recientes': reportes_recientes_data,
            }
            
            return Response(stats, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.warning(f"[WARNING] Error obteniendo estadsticas de reportes: {e}")
            # Retornar datos vacos en lugar de 500
            return Response({
                'total_reportes': 0,
                'reportes_completados': 0,
                'reportes_generando': 0,
                'reportes_fallidos': 0,
                'reportes_por_tipo': {},
                'reportes_por_formato': {},
                'reportes_recientes': []
            }, status=status.HTTP_200_OK)


class ReporteCleanupView(AdminPermissionMixin, APIView):
    """
    Vista para limpiar reportes expirados (solo administradores).
    """
    permission_classes = [IsAuthenticated]
    
    @swagger_auto_schema(
        operation_description="Limpia reportes expirados del sistema (solo administradores)",
        operation_summary="Limpiar reportes expirados",
        responses={
            200: openapi.Response(description="Limpieza completada exitosamente"),
            403: ErrorResponseSerializer,
            401: ErrorResponseSerializer,
        },
        tags=['Reportes']
    )
    def post(self, request):
        """Limpiar reportes expirados."""
        try:
            if not self.is_admin_user(request.user):
                return self.admin_permission_denied()
            
            # Limpiar reportes expirados
            cleaned_count = ReporteGenerado.limpiar_expirados()
            
            logger.info(f"Limpieza de reportes expirados completada por {request.user.username}: {cleaned_count} reportes eliminados")
            
            return Response({
                'message': f'Se limpiaron {cleaned_count} reportes expirados',
                'cleaned_count': cleaned_count,
                'status': 'success'
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.error(f"Error limpiando reportes expirados: {e}")
            return Response({
                'error': 'Error interno del servidor',
                'status': 'error'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class ReporteAgricultoresView(APIView):
    """
    Genera un archivo Excel con la informacin de agricultores y sus fincas.
    Requiere autenticacin JWT y permisos de administrador.
    """
    permission_classes = [IsAuthenticated, IsAdminUser]
    renderer_classes = [ExcelRenderer]
    
    def get(self, request, *args, **kwargs):
        """
        Genera y descarga un archivo Excel con informacin de todos los agricultores y sus fincas.
        """
        from io import BytesIO
        from openpyxl import Workbook
        from django.utils.encoding import escape_uri_path
        
        User = get_user_model()
        
        try:
            logger.info("[INFO] Generando reporte de agricultores para %s", request.user.username)
            
            # Validar que el modelo Finca est disponible
            if Finca is None:
                raise ImportError("El modelo Finca no est disponible")
            
            # Importar estilos de openpyxl
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
            
            # Crear libro y hoja de Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Agricultores"
            
            # Colores corporativos CacaoScan
            COLOR_VERDE_PRIMITIVO = "166534"
            COLOR_VERDE_CLARO = "A7F3D0"
            COLOR_GRIS_SUAVE = "6B7280"
            
            # Definir bordes delgados
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Ttulo principal (Fila 1)
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
            title_cell = ws.cell(row=1, column=1, value="Reporte de Agricultores y Fincas - CacaoScan")
            title_cell.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
            title_cell.fill = PatternFill(start_color=COLOR_VERDE_PRIMITIVO, end_color=COLOR_VERDE_PRIMITIVO, fill_type="solid")
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 30
            
            # Fila de separacin vaca (Fila 2)
            ws.row_dimensions[2].height = 10
            
            # Encabezados del Excel (Fila 3)
            headers = [
                'Agricultor', 'Email', 'Telefono', 'Departamento', 'Municipio',
                'Finca', 'Hectareas', 'Estado Finca', 'Fecha Registro Finca'
            ]
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=3, column=col_idx, value=header)
                cell.font = Font(name="Calibri", size=12, bold=True, color=COLOR_VERDE_PRIMITIVO)
                cell.fill = PatternFill(start_color=COLOR_VERDE_CLARO, end_color=COLOR_VERDE_CLARO, fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            ws.row_dimensions[3].height = 25
            
            # Obtener IDs de usuarios que tienen al menos una finca (agricultores)
            try:
                agricultores_ids = Finca.objects.values_list("agricultor_id", flat=True).distinct()
                agricultores_ids_list = list(agricultores_ids)
                
                logger.debug("[DEBUG] Encontrados %d agricultores con fincas", len(agricultores_ids_list))
                
                if not agricultores_ids_list:
                    logger.warning("[WARNING] No hay agricultores con fincas en la base de datos")
                    agricultores_list = []
                    # Agregar fila indicando que no hay datos (fila 4)
                    data_row = 4
                    no_data_row = ["-", "Sin datos", "-", "-", "-", "-", "-", "-", "-"]
                    for col_idx, value in enumerate(no_data_row, start=1):
                        cell = ws.cell(row=data_row, column=col_idx, value=value)
                        cell.font = Font(name="Calibri", size=11)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = thin_border
                else:
                    # Obtener los usuarios con prefetch_related para optimizar
                    agricultores = User.objects.filter(id__in=agricultores_ids_list).prefetch_related('fincas_app_fincas')
                    agricultores_list = list(agricultores)
                    logger.info("[INFO] Obtenidos %d agricultores con prefetch", len(agricultores_list))
                    
            except Exception as query_error:
                logger.error("[ERROR] Error obteniendo agricultores: %s", query_error, exc_info=True)
                raise
            
            # Procesar agricultores y agregar datos al Excel
            # Los datos empiezan en la fila 4 (despus del ttulo, separacin y encabezados)
            data_start_row = 4
            current_row = data_start_row
            rows_added = 0
            
            for agricultor in agricultores_list:
                try:
                    # Informacin bsica del agricultor
                    nombre = agricultor.get_full_name() or agricultor.username or f"Usuario {agricultor.id}"
                    email = agricultor.email or ""
                    
                    # Obtener telfono de forma segura
                    telefono = ""
                    try:
                        persona = getattr(agricultor, 'persona', None)
                        if persona and hasattr(persona, 'telefono') and persona.telefono:
                            telefono = str(persona.telefono)
                        elif hasattr(agricultor, 'auth_profile') and agricultor.auth_profile:
                            telefono = str(agricultor.auth_profile.phone_number) if agricultor.auth_profile.phone_number else ""
                    except Exception:
                        pass  # Ignorar errores de telfono (no crtico)
                    
                    # Obtener fincas del agricultor
                    fincas_list = []
                    try:
                        if hasattr(agricultor, 'fincas_app_fincas'):
                            fincas_queryset = agricultor.fincas_app_fincas.all()
                            fincas_list = list(fincas_queryset)
                    except Exception as finca_error:
                        logger.warning("[WARNING] Error obteniendo fincas para usuario %d: %s", agricultor.id, finca_error)
                        fincas_list = []
                    
                    # Si el agricultor tiene fincas, crear una fila por cada finca
                    if fincas_list:
                        for finca in fincas_list:
                            try:
                                # Validar y convertir valores de forma segura
                                depto = str(finca.departamento) if finca.departamento else ""
                                municipio = str(finca.municipio) if finca.municipio else ""
                                finca_nombre = str(finca.nombre) if finca.nombre else "Sin nombre"
                                
                                # Convertir hectreas
                                hectareas_val = 0.0
                                try:
                                    if finca.hectareas is not None:
                                        hectareas_val = float(finca.hectareas)
                                except (ValueError, TypeError, AttributeError):
                                    hectareas_val = 0.0
                                
                                # Estado de finca
                                estado_finca = "Activa" if (hasattr(finca, 'activa') and finca.activa) else "Inactiva"
                                
                                # Fecha de registro con formato dd/mm/yyyy
                                fecha_registro_str = ""
                                try:
                                    if hasattr(finca, 'fecha_registro') and finca.fecha_registro:
                                        fecha_registro_str = finca.fecha_registro.strftime('%d/%m/%Y')
                                except Exception:
                                    pass
                                
                                # Agregar fila al Excel con estilos
                                row_data = [
                                    nombre, email, telefono, depto, municipio,
                                    finca_nombre, hectareas_val, estado_finca, fecha_registro_str
                                ]
                                
                                for col_idx, value in enumerate(row_data, start=1):
                                    cell = ws.cell(row=current_row, column=col_idx, value=value)
                                    cell.font = Font(name="Calibri", size=11)
                                    cell.alignment = Alignment(horizontal="left", vertical="center")
                                    cell.border = thin_border
                                    
                                    # Nmeros (hectreas) alineados a la derecha
                                    if col_idx == 7 and isinstance(value, (int, float)):
                                        cell.alignment = Alignment(horizontal="right", vertical="center")
                                
                                current_row += 1
                                rows_added += 1
                                
                            except Exception as finca_row_error:
                                logger.warning("[WARNING] Error procesando finca %s: %s", 
                                             getattr(finca, 'id', 'unknown'), finca_row_error)
                                continue
                    else:
                        # Sin fincas, crear fila con datos del agricultor solamente
                        fecha_registro_str = ""
                        try:
                            if agricultor.date_joined:
                                fecha_registro_str = agricultor.date_joined.strftime('%d/%m/%Y')
                        except Exception:
                            pass
                        
                        row_data = [
                            nombre, email, telefono,
                            "", "", "", "", "",  # Departamento, Municipio, Finca, Hectreas, Estado
                            fecha_registro_str
                        ]
                        
                        for col_idx, value in enumerate(row_data, start=1):
                            cell = ws.cell(row=current_row, column=col_idx, value=value)
                            cell.font = Font(name="Calibri", size=11)
                            cell.alignment = Alignment(horizontal="left", vertical="center")
                            cell.border = thin_border
                        
                        current_row += 1
                        rows_added += 1
                        
                except Exception as agricultor_error:
                    logger.warning("[WARNING] Error procesando agricultor %s: %s", 
                                 getattr(agricultor, 'id', 'unknown'), agricultor_error)
                    continue
            
            logger.info("[INFO] Procesados %d filas de datos", rows_added)
            
            # Agregar pie de pgina
            footer_start_row = current_row + 2  # 2 filas vacas despus de los datos
            
            # Fila vaca 1
            ws.row_dimensions[footer_start_row - 1].height = 10
            
            # Generado por (fila siguiente)
            ws.merge_cells(start_row=footer_start_row, start_column=1, end_row=footer_start_row, end_column=9)
            generated_cell = ws.cell(row=footer_start_row, column=1, value=f"Generado por: {request.user.username}")
            generated_cell.font = Font(name="Calibri", size=10, color=COLOR_GRIS_SUAVE)
            generated_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Fecha de generacin (fila siguiente) - formato dd/mm/yyyy hh:mm:ss AM/PM (hora local)
            from datetime import datetime
            now = datetime.now()  # Usar hora local en lugar de UTC
            fecha_generacion = now.strftime('%d/%m/%Y %I:%M:%S %p')  # Formato 12 horas con AM/PM incluyendo segundos
            ws.merge_cells(start_row=footer_start_row + 1, start_column=1, end_row=footer_start_row + 1, end_column=9)
            fecha_cell = ws.cell(row=footer_start_row + 1, column=1, value=f"Fecha: {fecha_generacion}")
            fecha_cell.font = Font(name="Calibri", size=10, color=COLOR_GRIS_SUAVE)
            fecha_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Ajustar ancho de columnas automticamente
            column_widths = {
                'A': 25,  # Agricultor
                'B': 30,  # Email
                'C': 15,  # Telefono
                'D': 18,  # Departamento
                'E': 18,  # Municipio
                'F': 20,  # Finca
                'G': 12,  # Hectareas
                'H': 15,  # Estado Finca
                'I': 18,  # Fecha Registro
            }
            
            for col_letter, width in column_widths.items():
                ws.column_dimensions[col_letter].width = width
            
            # Guardar el archivo en memoria
            output = BytesIO()
            wb.save(output)
            output.seek(0)  # MUY IMPORTANTE: reposicionar el puntero al inicio del archivo
            
            # Obtener el contenido del buffer
            excel_content = output.getvalue()
            
            # Validar contenido generado
            if not excel_content or len(excel_content) < 50:
                output.close()
                logger.error("[ERROR] El archivo Excel generado est vaco o demasiado pequeo: %d bytes", 
                           len(excel_content) if excel_content else 0)
                return HttpResponse(
                    "Error interno al generar el reporte Excel: contenido invlido",
                    status=500,
                    content_type='text/plain'
                )
            
            logger.info("[SUCCESS] Reporte generado correctamente (%d bytes) para usuario %s", 
                       len(excel_content), request.user.username)
            
            # Nombre del archivo con formato seguro
            fecha_actual = timezone.now().strftime("%Y-%m-%d_%H-%M-%S")
            filename = f"reporte_agricultores_{fecha_actual}.xlsx"
            
            # Retornar el archivo correctamente formateado
            response = HttpResponse(
                excel_content,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = f'attachment; filename="{escape_uri_path(filename)}"'
            response["Content-Length"] = str(output.getbuffer().nbytes)
            
            # Cerrar el buffer correctamente
            output.close()
            
            # Retornar HttpResponse directamente - DRF lo detectar y no intentar serializarlo
            return response
            
        except Exception as e:
            error_trace = traceback.format_exc()
            logger.error("[ERROR] Error generando reporte de agricultores: %s\n%s", str(e), error_trace)
            
            return HttpResponse(
                f"Error interno al generar el reporte: {str(e)}",
                status=500,
                content_type='text/plain'
            )


class ReporteUsuariosView(APIView):
    """
    Vista para generar y descargar reporte Excel de usuarios del sistema.
    Solo accesible para administradores.
    """
    permission_classes = [IsAuthenticated, IsAdminUser]
    # Deshabilitar renderers de DRF para devolver respuesta binaria directamente
    renderer_classes = []
    
    @swagger_auto_schema(
        operation_description="Genera y descarga un archivo Excel con información de todos los usuarios del sistema",
        operation_summary="Reporte de usuarios (Excel)",
        responses={
            200: openapi.Response(description="Archivo Excel generado exitosamente"),
            401: ErrorResponseSerializer,
            403: ErrorResponseSerializer,
        },
        tags=['Reportes']
    )
    def get(self, request):
        try:
            # Generar el reporte Excel
            excel_service = ExcelUsuariosService()
            excel_content = excel_service.generate_users_report()
            
            # Validar que el contenido no esté vacío
            if not excel_content or len(excel_content) < 100:
                logger.error("El contenido del reporte Excel está vacío o corrupto")
                return HttpResponse(
                    json.dumps({'error': 'Error al generar el reporte Excel: contenido vacío'}),
                    content_type='application/json',
                    status=500
                )
            
            # Crear respuesta HTTP con el archivo (usar HttpResponse directamente, no Response de DRF)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'reporte_usuarios_{timestamp}.xlsx'
            
            response = HttpResponse(
                excel_content,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
            # Configurar nombre del archivo con encoding correcto
            from django.utils.encoding import escape_uri_path
            response['Content-Disposition'] = f'attachment; filename="{escape_uri_path(filename)}"'
            response['Content-Length'] = len(excel_content)
            
            logger.info(f"Reporte de usuarios generado por {request.user.username} - Tamaño: {len(excel_content)} bytes")
            
            return response
            
        except Exception as e:
            logger.error(f"Error generando reporte de usuarios: {e}", exc_info=True)
            # Para errores, usar HttpResponse con JSON en lugar de Response de DRF
            error_response = HttpResponse(
                json.dumps({
                    'error': 'Error al generar el reporte de usuarios',
                    'detail': str(e)
                }),
                content_type='application/json',
                status=500
            )
            return error_response


